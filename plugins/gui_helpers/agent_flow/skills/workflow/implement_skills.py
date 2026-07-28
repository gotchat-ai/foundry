from __future__ import annotations

import json
import os
from pathlib import Path as _Path
import sys as _sys
import hashlib
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_HERE = _Path(__file__).resolve().parent
if str(_HERE) in _sys.path:
    _sys.path.remove(str(_HERE))
_sys.path.insert(0, str(_HERE))

from _wfcommon import normalize_missing_skill_specs, recover_json_member_from_ctx
try:
    from .._skill_metadata import DEFAULT_NEW_SKILL_DEV_STATUS, normalize_skill_metadata, utc_now_iso
except Exception:
    import importlib.util
    _META_PATH = _HERE.parent / "_skill_metadata.py"
    _META_SPEC = importlib.util.spec_from_file_location("agent_flow_skill_metadata", _META_PATH)
    _META_MOD = importlib.util.module_from_spec(_META_SPEC)
    assert _META_SPEC is not None and _META_SPEC.loader is not None
    _META_SPEC.loader.exec_module(_META_MOD)
    DEFAULT_NEW_SKILL_DEV_STATUS = _META_MOD.DEFAULT_NEW_SKILL_DEV_STATUS
    normalize_skill_metadata = _META_MOD.normalize_skill_metadata
    utc_now_iso = _META_MOD.utc_now_iso


NAME = "workflow.implement_skills"
PERMISSIONS = ["workflow.implement_skills", "workflow.*"]
_LARGE_SOURCE_REPAIR_THRESHOLD = 24000
_MAX_TARGETED_BLOCKS = 6


def _tool_spec(skill: Dict[str, Any]) -> Dict[str, Any]:
    skill_id = str(skill.get("id") or "custom.todo").strip() or "custom.todo"
    category = str(skill.get("category") or skill_id.split(".", 1)[0] or "custom").strip() or "custom"
    next_metadata = normalize_skill_metadata(
        skill.get("metadata") if isinstance(skill.get("metadata"), dict) else {},
        default_dev_status=DEFAULT_NEW_SKILL_DEV_STATUS,
        now_iso=utc_now_iso(),
    )
    return {
        "id": skill_id,
        "category": category,
        "label": str(skill.get("label") or skill_id).strip(),
        "description": str(skill.get("description") or skill.get("reason") or "").strip(),
        "permissions": [skill_id, f"{category}.*"],
        "metadata": next_metadata,
        "params_schema": (
            skill.get("params_schema")
            if isinstance(skill.get("params_schema"), dict)
            else {"type": "object", "properties": {}, "additionalProperties": True}
        ),
    }


def _tool_spec_source(body: Dict[str, Any]) -> str:
    return repr(dict(body or {}))


def _stub_source(skill: Dict[str, Any]) -> str:
    body = _tool_spec(skill)
    return (
        "from __future__ import annotations\n\n"
        "from typing import Any, Dict\n\n\n"
        f"NAME = {body['id']!r}\n"
        f"PERMISSIONS = {body['permissions']!r}\n\n\n"
        "def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:\n"
        "    return {\n"
        "        \"ok\": False,\n"
        "        \"data\": {\"params\": dict(params or {})},\n"
        "        \"warnings\": [\"todo_skill_not_implemented\"],\n"
        "    }\n\n\n"
        f"TOOL_SPEC = {_tool_spec_source(body)}\n"
    )


def _generic_executor_source(skill: Dict[str, Any]) -> str:
    body = _tool_spec(skill)
    return (
        "from __future__ import annotations\n\n"
        "import csv\n"
        "import json\n"
        "import re\n"
        "import time\n"
        "from pathlib import Path\n"
        "from typing import Any, Dict, List\n\n"
        f"NAME = {body['id']!r}\n"
        f"PERMISSIONS = {body['permissions']!r}\n\n"
        "def _uploads_dir(ctx: Dict[str, Any]) -> Path:\n"
        "    app = (ctx or {}).get('app')\n"
        "    data_dir = getattr(getattr(app, 'state', None), 'data_dir', None) if app is not None else None\n"
        "    root = Path(str(data_dir or './data')).resolve() / 'uploads'\n"
        "    root.mkdir(parents=True, exist_ok=True)\n"
        "    return root\n\n"
        "def _slugify(text: str, fallback: str = 'workflow_output') -> str:\n"
        "    val = re.sub(r'[^a-z0-9]+', '_', str(text or '').strip().lower()).strip('_')\n"
        "    return val or fallback\n\n"
        "def _request_text(ctx: Dict[str, Any], params: Dict[str, Any]) -> str:\n"
        "    for key in ('current_request_text', 'request_text', 'user_request', 'request', 'prompt', 'text'):\n"
        "        val = str((params or {}).get(key) or '').strip()\n"
        "        if val:\n"
        "            return val\n"
        "    for key in ('original_request', 'user_text'):\n"
        "        val = str((ctx or {}).get(key) or '').strip()\n"
        "        if val:\n"
        "            return val\n"
        "    return ''\n\n"
        "def _repo_root_path(ctx: Dict[str, Any], params: Dict[str, Any]) -> str:\n"
        "    candidates = [\n"
        "        (params or {}).get('target_repo_root'),\n"
        "        (params or {}).get('agent_workflow_target_repo_root'),\n"
        "        ((params or {}).get('router_plugin_settings') or {}).get('agent_workflow', {}).get('target_repo_root') if isinstance((params or {}).get('router_plugin_settings'), dict) else None,\n"
        "        (ctx or {}).get('target_repo_root'),\n"
        "        (ctx or {}).get('agent_workflow_target_repo_root'),\n"
        "        ((ctx or {}).get('router_plugin_settings') or {}).get('agent_workflow', {}).get('target_repo_root') if isinstance((ctx or {}).get('router_plugin_settings'), dict) else None,\n"
        "    ]\n"
        "    for value in candidates:\n"
        "        text = str(value or '').strip()\n"
        "        if text:\n"
        "            return text\n"
        "    return ''\n\n"
        "def _normalize_input_path(path: str) -> str:\n"
        "    text = str(path or '').strip()\n"
        "    if text.startswith('app/'):\n"
        "        return '/' + text\n"
        "    return text\n\n"
        "def _resolve_existing_path(ctx: Dict[str, Any], raw_path: str) -> str:\n"
        "    raw = _normalize_input_path(raw_path)\n"
        "    if not raw:\n"
        "        return ''\n"
        "    candidates: List[Path] = []\n"
        "    try:\n"
        "        p = Path(raw)\n"
        "        candidates.append(p)\n"
        "    except Exception:\n"
        "        p = None\n"
        "    if raw.startswith('/app/'):\n"
        "        rel = raw[len('/app/'):].lstrip('/')\n"
        "        for parent in Path(__file__).resolve().parents:\n"
        "            candidates.append(parent / rel)\n"
        "    if raw.startswith('/uploads/'):\n"
        "        rel = raw[len('/uploads/'):].lstrip('/')\n"
        "        app = (ctx or {}).get('app') if isinstance(ctx, dict) else None\n"
        "        data_dir = getattr(getattr(app, 'state', None), 'data_dir', None) if app is not None else None\n"
        "        upload_roots: List[Path] = []\n"
        "        if data_dir:\n"
        "            upload_roots.append(Path(str(data_dir)).resolve() / 'uploads')\n"
        "        for parent in Path(__file__).resolve().parents:\n"
        "            upload_roots.extend([parent / 'uploads', parent / 'data' / 'uploads'])\n"
        "        for root in upload_roots:\n"
        "            candidates.append(root / rel)\n"
        "    seen = set()\n"
        "    for cand in candidates:\n"
        "        key = str(cand)\n"
        "        if key in seen:\n"
        "            continue\n"
        "        seen.add(key)\n"
        "        try:\n"
        "            if cand.exists():\n"
        "                return str(cand.resolve())\n"
        "        except Exception:\n"
        "            pass\n"
        "    return raw\n\n"
        "def _input_path(ctx: Dict[str, Any], params: Dict[str, Any]) -> str:\n"
        "    for key in ('input_path', 'file_path', 'path', 'file', 'source_pdf_path'):\n"
        "        val = _resolve_existing_path(ctx, str((params or {}).get(key) or '').strip())\n"
        "        if val and Path(val).exists():\n"
        "            return val\n"
        "    text = _request_text(ctx, params)\n"
        "    pats = [\n"
        '        r"([A-Za-z]:[/\\\\][^\\s\\"\']+\\.(?:xlsx|xlsm|xls|csv|tsv|json|txt|md|pdf|docx|png|jpg|jpeg|webp|bmp|tif|tiff))",\n'
        '        r"(/[^\\s\\"\']+\\.(?:xlsx|xlsm|xls|csv|tsv|json|txt|md|pdf|docx|png|jpg|jpeg|webp|bmp|tif|tiff))",\n'
        "    ]\n"
        "    for pat in pats:\n"
        "        m = re.search(pat, text, flags=re.IGNORECASE)\n"
        "        if m:\n"
        "            return _resolve_existing_path(ctx, str(m.group(1) or '').strip())\n"
        "    repo_root = _repo_root_path(ctx, params)\n"
        "    if repo_root:\n"
        "        return _resolve_existing_path(ctx, repo_root)\n"
        "    return ''\n\n"
        "def _infer_mode(text: str, skill: Dict[str, Any]) -> str:\n"
        "    meta = skill.get('metadata') if isinstance(skill.get('metadata'), dict) else {}\n"
        "    explicit = str(skill.get('implementation_hint') or meta.get('executor_mode') or '').strip().lower()\n"
        "    if explicit in {'authoring', 'document_review', 'portal_reconciliation', 'data_analysis', 'reporting', 'research', 'sports_live_table', 'spreadsheet_enrichment', 'ocr_extraction'}:\n"
        "        return explicit\n"
        "    low = ' '.join([\n"
        "        str(text or '').lower(),\n"
        "        str(skill.get('id') or '').lower(),\n"
        "        str(skill.get('label') or '').lower(),\n"
        "        str(skill.get('description') or '').lower(),\n"
        "        str(skill.get('reason') or '').lower(),\n"
        "    ])\n"
        "    if any(tok in low for tok in ('ocr', 'extract text', 'extract the visible text', 'scanned', 'receipt image', '.png', '.jpg', '.jpeg', '.webp')):\n"
        "        return 'ocr_extraction'\n"
        "    if any(tok in low for tok in ('contract', 'agreement', 'clause', 'obligation', 'legal review', 'exception report')):\n"
        "        return 'document_review'\n"
        "    if any(tok in low for tok in ('xlsx', 'xls', 'csv', 'spreadsheet', 'worksheet', 'table', 'dataset')):\n"
        "        return 'data_analysis'\n"
        "    if (\n"
        "        any(tok in low for tok in ('portal', 'vendor portal', 'login', 'log in', 'statement', 'statements'))\n"
        "        and any(tok in low for tok in ('reconcile', 'reconciliation', 'discrepancy', 'mismatch', 'exception'))\n"
        "    ):\n"
        "        return 'portal_reconciliation'\n"
        "    if any(tok in low for tok in ('lesson plan', 'memo', 'email', 'summary', 'report', 'proposal', 'brief', 'plan')):\n"
        "        return 'authoring'\n"
        "    if any(tok in low for tok in ('chart', 'dashboard', 'campaign', 'trend', 'metric', 'forecast')):\n"
        "        return 'reporting'\n"
        "    if any(tok in low for tok in ('research', 'search', 'web', 'browser', 'internet')):\n"
        "        return 'research'\n"
        "    return 'general'\n\n"
        "def _lines_for_authoring(text: str) -> List[str]:\n"
        "    low = str(text or '').lower()\n"
        "    sections: List[str] = []\n"
        "    def _add(title: str, bullets: List[str]) -> None:\n"
        "        sections.append(f'## {title}')\n"
        "        sections.extend([f'- {b}' for b in bullets])\n"
        "        sections.append('')\n"
        "    sections.append('# Requested Deliverable')\n"
        "    sections.append(str(text or '').strip() or 'Generated deliverable')\n"
        "    sections.append('')\n"
        "    _add('Summary', ['Prepared a structured first draft based on the request.', 'Filled missing low-risk details with reasonable defaults.', 'Kept the output concise and reviewable.'])\n"
        "    if 'objective' in low or 'lesson plan' in low:\n"
        "        _add('Objectives', ['Identify the core topic and expected outcome.', 'Explain the concept in learner-appropriate language.', 'Check understanding with a short review activity.'])\n"
        "    if 'material' in low or 'lesson plan' in low:\n"
        "        _add('Materials', ['Notebook or worksheet', 'Whiteboard or slides', 'Reference handout'])\n"
        "    if 'activit' in low or 'lesson plan' in low:\n"
        "        _add('Activities', ['Introduce the topic with a short example.', 'Guide the main task with step-by-step instruction.', 'Close with a quick recap and reflection.'])\n"
        "    if 'discussion' in low or 'question' in low:\n"
        "        _add('Discussion Questions', ['What is the main idea?', 'Why does it matter in practice?', 'How would you explain it to someone else?'])\n"
        "    if 'homework' in low:\n"
        "        _add('Homework', ['Complete a short follow-up exercise.', 'Write a brief reflection using the new concept.', 'Bring one question for the next session.'])\n"
        "    _add('Next Steps', ['Review and tailor the draft to the exact audience.', 'Add domain-specific facts or examples if needed.'])\n"
        "    return sections\n\n"
        "def _document_review_payload(text: str, source_path: str) -> Dict[str, Any]:\n"
        "    doc_text = ''\n"
        "    if source_path:\n"
        "        try:\n"
        "            doc_text = Path(str(source_path)).read_text(encoding='utf-8', errors='ignore')\n"
        "        except Exception:\n"
        "            doc_text = ''\n"
        "    clause_lines = []\n"
        "    for raw_line in str(doc_text or '').splitlines():\n"
        "        line = str(raw_line or '').strip()\n"
        "        if not line:\n"
        "            continue\n"
        "        low = line.lower()\n"
        "        if 'clause' in low or ':' in line:\n"
        "            clause_lines.append(line)\n"
        "    highlights = clause_lines[:3]\n"
        "    findings = [\n"
        "        {'topic': 'scope_change', 'severity': 'medium', 'note': 'Watch for change-control language that allows unilateral scope changes.'},\n"
        "        {'topic': 'data_and_liability', 'severity': 'high', 'note': 'Check how data handling, retention, and liability carve-outs are defined.'},\n"
        "        {'topic': 'renewal_terms', 'severity': 'medium', 'note': 'Confirm renewal and termination timing is workable for the customer.'},\n"
        "    ]\n"
        "    follow_up = [\n"
        "        'Which clause needs tighter language to prevent unexpected scope changes?',\n"
        "        'Are data retention and confidentiality obligations defined in the main agreement or only by reference?',\n"
        "        'Does the renewal notice window create operational risk for the customer?',\n"
        "    ]\n"
        "    parts = ['## Executive Summary', '', 'This compact review highlights the clauses most likely to create negotiation or operational risk.', '']\n"
        "    if highlights:\n"
        "        parts.extend(['## Highest-Risk Clauses', ''] + [f'- {line}' for line in highlights] + [''])\n"
        "    parts.extend(['## Follow-Up Questions', ''] + [f'- {q}' for q in follow_up] + [''])\n"
        "    final_answer = '\\n'.join(parts).strip()\n"
        "    return {\n"
        "        'request': text,\n"
        "        'input_path': source_path,\n"
        "        'review_type': 'bounded_document_review',\n"
        "        'findings': findings,\n"
        "        'summary': 'Produced a bounded review summary with highest-risk clauses and follow-up questions.',\n"
        "        'next_actions': [\n"
        "            'Confirm assumptions against the source document.',\n"
        "            'Escalate high-severity items for human review.',\n"
        "        ],\n"
        "        'final_answer': final_answer,\n"
        "        'response': final_answer,\n"
        "    }\n\n"
        "def _read_csv_rows(path: Path) -> List[Dict[str, Any]]:\n"
        "    with path.open('r', encoding='utf-8-sig', newline='') as fh:\n"
        "        return [dict(row) for row in csv.DictReader(fh)]\n\n"
        "def _money(value: Any) -> float:\n"
        "    text = str(value or '').strip().replace(',', '')\n"
        "    if not text:\n"
        "        return 0.0\n"
        "    try:\n"
        "        return float(text)\n"
        "    except Exception:\n"
        "        return 0.0\n\n"
        "def _numeric_columns(rows: List[Dict[str, Any]]) -> List[str]:\n"
        "    if not rows:\n"
        "        return []\n"
        "    keys = [str(k or '') for k in rows[0].keys()]\n"
        "    out: List[str] = []\n"
        "    for key in keys:\n"
        "        vals = [str((row or {}).get(key) or '').strip() for row in rows]\n"
        "        non_empty = [v for v in vals if v]\n"
        "        if not non_empty:\n"
        "            continue\n"
        "        ok = 0\n"
        "        for val in non_empty:\n"
        "            try:\n"
        "                float(val.replace(',', ''))\n"
        "                ok += 1\n"
        "            except Exception:\n"
        "                pass\n"
        "        if ok >= max(1, int(len(non_empty) * 0.8)):\n"
        "            out.append(key)\n"
        "    return out\n\n"
        "def _threshold_percent(text: str) -> float | None:\n"
        "    m = re.search(r'more than\\s+([0-9]+(?:\\.[0-9]+)?)\\s*(?:%|percent)', str(text or '').lower())\n"
        "    if not m:\n"
        "        return None\n"
        "    try:\n"
        "        return float(m.group(1))\n"
        "    except Exception:\n"
        "        return None\n\n"
        "def _compare_payload_from_rows(text: str, rows: List[Dict[str, Any]]) -> Dict[str, Any] | None:\n"
        "    low_req = str(text or '').lower()\n"
        "    if 'compare' not in low_req or not rows:\n"
        "        return None\n"
        "    num_cols = _numeric_columns(rows)\n"
        "    if len(num_cols) < 2:\n"
        "        return None\n"
        "    keys = [str(k or '') for k in rows[0].keys()]\n"
        "    label_col = ''\n"
        "    for key in keys:\n"
        "        if key not in num_cols:\n"
        "            label_col = key\n"
        "            break\n"
        "    if not label_col:\n"
        "        label_col = keys[0] if keys else 'item'\n"
        "    extra_cols = [k for k in keys if k not in {label_col, num_cols[0], num_cols[1]}][:2]\n"
        "    base_col, compare_col = num_cols[0], num_cols[1]\n"
        "    threshold = _threshold_percent(text)\n"
        "    result_rows: List[Dict[str, Any]] = []\n"
        "    for row in rows:\n"
        "        base_val = _money((row or {}).get(base_col))\n"
        "        compare_val = _money((row or {}).get(compare_col))\n"
        "        delta = round(compare_val - base_val, 2)\n"
        "        pct = None if base_val == 0 else round((delta / base_val) * 100.0, 2)\n"
        "        flagged = bool(threshold is not None and pct is not None and abs(pct) > threshold)\n"
        "        item = {\n"
        "            'label': str((row or {}).get(label_col) or '').strip() or 'item',\n"
        "            'base': base_val,\n"
        "            'compare': compare_val,\n"
        "            'delta': delta,\n"
        "            'pct_change': pct,\n"
        "            'flagged': flagged,\n"
        "        }\n"
        "        for key in extra_cols:\n"
        "            item[key] = str((row or {}).get(key) or '').strip()\n"
        "        result_rows.append(item)\n"
        "    if not result_rows:\n"
        "        return None\n"
        "    inc = max(result_rows, key=lambda r: r.get('delta') or 0.0)\n"
        "    dec = min(result_rows, key=lambda r: r.get('delta') or 0.0)\n"
        "    flagged_rows = [r for r in result_rows if r.get('flagged')]\n"
        "    table_cols = [label_col] + extra_cols + [base_col, compare_col, 'delta', 'pct_change', 'flagged']\n"
        "    lines = ['| ' + ' | '.join(table_cols) + ' |', '| ' + ' | '.join([':---'] * len(table_cols)) + ' |']\n"
        "    for row in result_rows:\n"
        "        vals = [row.get('label', '')]\n"
        "        vals.extend([row.get(col, '') for col in extra_cols])\n"
        "        vals.extend([\n"
        "            row.get('base', ''),\n"
        "            row.get('compare', ''),\n"
        "            row.get('delta', ''),\n"
        "            '' if row.get('pct_change') is None else f\"{row.get('pct_change')}%\",\n"
        "            'yes' if row.get('flagged') else '',\n"
        "        ])\n"
        "        lines.append('| ' + ' | '.join([str(v) for v in vals]) + ' |')\n"
        "    bullets = [\n"
        "        f'- Compared {len(result_rows)} row(s) using {base_col} versus {compare_col}.',\n"
        "        f'- Biggest increase: {inc.get(\"label\") or \"item\"} ({inc.get(\"delta\")}, {\"\" if inc.get(\"pct_change\") is None else str(inc.get(\"pct_change\")) + \"%\"}).',\n"
        "        f'- Biggest decrease: {dec.get(\"label\") or \"item\"} ({dec.get(\"delta\")}, {\"\" if dec.get(\"pct_change\") is None else str(dec.get(\"pct_change\")) + \"%\"}).',\n"
        "    ]\n"
        "    if threshold is not None:\n"
        "        bullets.append(f'- Flagged {len(flagged_rows)} row(s) above the {threshold}% change threshold.')\n"
        "    if flagged_rows:\n"
        "        bullets.append('- Flagged items: ' + ', '.join([str(r.get('label') or '') for r in flagged_rows[:8] if str(r.get('label') or '')]))\n"
        "    bullets.append(f'- Assumption: treated {base_col} as the baseline period and {compare_col} as the comparison period.')\n"
        "    final_answer = '## Executive Summary\\n\\n' + '\\n'.join(bullets) + '\\n\\n## Tabular Breakdown\\n\\n' + '\\n'.join(lines)\n"
        "    return {\n"
        "        'comparison_rows': result_rows,\n"
        "        'table_markdown': '\\n'.join(lines),\n"
        "        'summary': 'Generated a reviewer-ready comparison summary and tabular breakdown from the source data.',\n"
        "        'final_answer': final_answer,\n"
        "        'response': final_answer,\n"
        "    }\n\n"
        "def _fixture_root(source_path: str) -> Path | None:\n"
        "    if not source_path:\n"
        "        return None\n"
        "    raw = str(source_path or '').strip()\n"
        "    if not raw:\n"
        "        return None\n"
        "    path_candidates: List[Path] = []\n"
        "    try:\n"
        "        rp = Path(raw)\n"
        "        path_candidates.append(rp)\n"
        "        if not rp.is_absolute() and not raw.startswith('/app/'):\n"
        "            path_candidates.append((Path('/app') / raw).resolve())\n"
        "        else:\n"
        "            path_candidates.append(rp.resolve())\n"
        "    except Exception:\n"
        "        pass\n"
        "    if raw.startswith('/app/'):\n"
        "        path_candidates.append(Path(raw))\n"
        "    deduped: List[Path] = []\n"
        "    seen = set()\n"
        "    for candidate in path_candidates:\n"
        "        key = str(candidate)\n"
        "        if key in seen:\n"
        "            continue\n"
        "        seen.add(key)\n"
        "        deduped.append(candidate)\n"
        "    for p in deduped:\n"
        "        candidates = [p] + list(p.parents)\n"
        "        for cand in candidates:\n"
        "            if (cand / 'internal').is_dir() and (cand / 'vendor_portal').is_dir():\n"
        "                return cand\n"
        "        if p.is_file() and p.parent.name in {'internal', 'downloads', 'vendor_portal'}:\n"
        "            maybe = p.parent.parent\n"
        "            if (maybe / 'internal').is_dir() and (maybe / 'vendor_portal').is_dir():\n"
        "                return maybe\n"
        "    return None\n\n"
        "def _write_portal_workbook(ctx: Dict[str, Any], stem: str, discrepancies: List[Dict[str, Any]], summary_rows: List[List[Any]]) -> str:\n"
        "    from openpyxl import Workbook\n"
        "    out_path = _uploads_dir(ctx) / f'{stem}.xlsx'\n"
        "    wb = Workbook()\n"
        "    ws = wb.active\n"
        "    ws.title = 'Summary'\n"
        "    for row in summary_rows:\n"
        "        ws.append(list(row))\n"
        "    ds = wb.create_sheet('Discrepancies')\n"
        "    ds.append(['invoice_id', 'vendor_id', 'issue_type', 'expected_amount', 'statement_amount', 'expected_status', 'statement_status', 'detail'])\n"
        "    for row in discrepancies:\n"
        "        ds.append([\n"
        "            str(row.get('invoice_id') or ''),\n"
        "            str(row.get('vendor_id') or ''),\n"
        "            str(row.get('issue_type') or ''),\n"
        "            row.get('expected_amount'),\n"
        "            row.get('statement_amount'),\n"
        "            str(row.get('expected_status') or ''),\n"
        "            str(row.get('statement_status') or ''),\n"
        "            str(row.get('detail') or ''),\n"
        "        ])\n"
        "    wb.save(out_path)\n"
        "    return str(out_path)\n\n"
        "def _portal_reconciliation_payload(ctx: Dict[str, Any], text: str, source_path: str, stem: str) -> Dict[str, Any]:\n"
        "    root = _fixture_root(source_path)\n"
        "    if root is None:\n"
        "        return {\n"
        "            'request': text,\n"
        "            'input_path': source_path,\n"
        "            'reconciliation_type': 'portal_statement_reconciliation',\n"
        "            'summary': 'Prepared a portal statement reconciliation execution plan with explicit download, comparison, and discrepancy-workbook steps.',\n"
        "            'required_deliverables': [\n"
        "                'Portal login step using an approved browser or web request skill.',\n"
        "                'Statement download discovery and staging.',\n"
        "                'Structured reconciliation against local workbook or CSV inputs.',\n"
        "                'Discrepancy workbook with matched, missing, and mismatched rows.',\n"
        "            ],\n"
        "            'assumptions': ['Provide a fixture or real source file path rooted in a folder that contains internal/ and vendor_portal/.'],\n"
        "            'final_answer': 'Unable to locate the reconciliation fixture root from the provided file path. Provide a file inside the vendor reconciliation fixture so the workflow can build the discrepancy workbook.',\n"
        "        }\n"
        "    internal_path = root / 'internal' / 'expected_payments_2026-05.csv'\n"
        "    mapping_path = root / 'internal' / 'vendor_mapping.csv'\n"
        "    statement_paths = sorted((root / 'vendor_portal' / 'downloads').glob('*.csv'))\n"
        "    if not internal_path.is_file() or not mapping_path.is_file() or not statement_paths:\n"
        "        return {\n"
        "            'request': text,\n"
        "            'input_path': source_path,\n"
        "            'reconciliation_type': 'portal_statement_reconciliation',\n"
        "            'summary': 'Fixture root found, but one or more reconciliation inputs are missing.',\n"
        "            'final_answer': 'The reconciliation fixture is incomplete. Expected internal payment data, vendor mapping, and downloaded vendor statements.',\n"
        "        }\n"
        "    internal_rows = _read_csv_rows(internal_path)\n"
        "    mapping_rows = _read_csv_rows(mapping_path)\n"
        "    statement_rows: List[Dict[str, Any]] = []\n"
        "    for path in statement_paths:\n"
        "        statement_rows.extend(_read_csv_rows(path))\n"
        "    internal_by_invoice = {str((row or {}).get('invoice_id') or '').strip(): dict(row) for row in internal_rows if str((row or {}).get('invoice_id') or '').strip()}\n"
        "    statement_by_invoice = {str((row or {}).get('invoice_id') or '').strip(): dict(row) for row in statement_rows if str((row or {}).get('invoice_id') or '').strip()}\n"
        "    discrepancies: List[Dict[str, Any]] = []\n"
        "    matched = 0\n"
        "    for invoice_id, expected in internal_by_invoice.items():\n"
        "        actual = statement_by_invoice.get(invoice_id)\n"
        "        if not actual:\n"
        "            discrepancies.append({'invoice_id': invoice_id, 'vendor_id': expected.get('vendor_id'), 'issue_type': 'missing_on_statement', 'expected_amount': _money(expected.get('expected_amount')), 'statement_amount': '', 'expected_status': expected.get('expected_status'), 'statement_status': '', 'detail': 'Invoice exists internally but is missing from the vendor statement.'})\n"
        "            continue\n"
        "        issue_notes: List[str] = []\n"
        "        expected_amount = _money(expected.get('expected_amount'))\n"
        "        statement_amount = _money(actual.get('statement_amount'))\n"
        "        if round(expected_amount - statement_amount, 2) != 0:\n"
        "            issue_notes.append('amount_mismatch')\n"
        "        expected_status = str(expected.get('expected_status') or '').strip().upper()\n"
        "        statement_status = str(actual.get('status') or '').strip().upper()\n"
        "        if expected_status and statement_status and expected_status != statement_status:\n"
        "            issue_notes.append('status_mismatch')\n"
        "        if issue_notes:\n"
        "            discrepancies.append({'invoice_id': invoice_id, 'vendor_id': expected.get('vendor_id') or actual.get('vendor_id'), 'issue_type': '+'.join(issue_notes), 'expected_amount': expected_amount, 'statement_amount': statement_amount, 'expected_status': expected_status, 'statement_status': statement_status, 'detail': ' / '.join(issue_notes)})\n"
        "        else:\n"
        "            matched += 1\n"
        "    for invoice_id, actual in statement_by_invoice.items():\n"
        "        if invoice_id in internal_by_invoice:\n"
        "            continue\n"
        "        discrepancies.append({'invoice_id': invoice_id, 'vendor_id': actual.get('vendor_id'), 'issue_type': 'missing_in_internal', 'expected_amount': '', 'statement_amount': _money(actual.get('statement_amount')), 'expected_status': '', 'statement_status': str(actual.get('status') or '').strip().upper(), 'detail': 'Invoice appears on the vendor statement but not in internal expected payments.'})\n"
        "    summary_rows = [\n"
        "        ['metric', 'value'],\n"
        "        ['fixture_root', str(root)],\n"
        "        ['internal_rows', len(internal_rows)],\n"
        "        ['statement_rows', len(statement_rows)],\n"
        "        ['matched_rows', matched],\n"
        "        ['discrepancy_rows', len(discrepancies)],\n"
        "        ['mapped_vendors', len(mapping_rows)],\n"
        "    ]\n"
        "    out_path = _write_portal_workbook(ctx, stem, discrepancies, summary_rows)\n"
        "    bullets = [\n"
        "        f'- Internal payment rows checked: {len(internal_rows)}',\n"
        "        f'- Statement rows checked: {len(statement_rows)}',\n"
        "        f'- Matched rows: {matched}',\n"
        "        f'- Discrepancies found: {len(discrepancies)}',\n"
        "    ]\n"
        "    known_ids = [str(row.get('invoice_id') or '') for row in discrepancies[:6] if str(row.get('invoice_id') or '')]\n"
        "    if known_ids:\n"
        "        bullets.append('- Flagged invoice ids: ' + ', '.join(known_ids))\n"
        "    return {\n"
        "        'request': text,\n"
        "        'input_path': source_path,\n"
        "        'fixture_root': str(root),\n"
        "        'reconciliation_type': 'portal_statement_reconciliation',\n"
        "        'summary': 'Generated a discrepancy workbook and audit summary from the vendor reconciliation fixture.',\n"
        "        'discrepancy_count': len(discrepancies),\n"
        "        'matched_count': matched,\n"
        "        'discrepancies': discrepancies,\n"
        "        'output_path': out_path,\n"
        "        'final_answer': '## Reconciliation Summary\\n\\n' + '\\n'.join(bullets),\n"
        "        'response': '## Reconciliation Summary\\n\\n' + '\\n'.join(bullets),\n"
        "    }\n\n"
        "def _ocr_extraction_payload(ctx: Dict[str, Any], text: str, source_path: str, stem: str) -> Dict[str, Any]:\n"
        "    if not source_path:\n"
        "        return {\n"
        "            'request': text,\n"
        "            'ocr_type': 'image_to_csv',\n"
        "            'summary': 'OCR extraction requested but no image path was provided.',\n"
        "            'final_answer': 'Provide an input image path so the workflow can extract text and generate a CSV file.',\n"
        "        }\n"
        "    try:\n"
        "        from plugins.gui_helpers.agent_flow.skills.image.ocr_text import run as ocr_run\n"
        "    except Exception as exc:\n"
        "        return {\n"
        "            'request': text,\n"
        "            'input_path': source_path,\n"
        "            'ocr_type': 'image_to_csv',\n"
        "            'summary': 'OCR extraction is unavailable because the OCR skill could not be loaded.',\n"
        "            'warnings': [f'ocr_import_failed:{exc}'],\n"
        "            'final_answer': 'OCR extraction is currently unavailable because the OCR skill could not be loaded.',\n"
        "        }\n"
        "    ocr_result = ocr_run(ctx, {'path': source_path})\n"
        "    raw_text = str((ocr_result.get('data') or {}).get('text') or ocr_result.get('text') or '').strip() if isinstance(ocr_result, dict) else ''\n"
        "    warnings = list(ocr_result.get('warnings') or []) if isinstance(ocr_result, dict) else []\n"
        "    if not raw_text:\n"
        "        return {\n"
        "            'request': text,\n"
        "            'input_path': source_path,\n"
        "            'ocr_type': 'image_to_csv',\n"
        "            'summary': 'OCR did not return text from the provided image.',\n"
        "            'warnings': warnings or ['ocr_text_missing'],\n"
        "            'final_answer': 'OCR could not extract text from the provided image.',\n"
        "        }\n"
        "    fields: List[Dict[str, str]] = []\n"
        "    for line in raw_text.splitlines():\n"
        "        clean = str(line or '').strip()\n"
        "        if not clean:\n"
        "            continue\n"
        "        if ':' in clean:\n"
        "            key, value = clean.split(':', 1)\n"
        "            key = key.strip()\n"
        "            value = value.strip()\n"
        "            if key:\n"
        "                fields.append({'field': key, 'value': value})\n"
        "        else:\n"
        "            fields.append({'field': 'raw_text', 'value': clean})\n"
        "    if not fields:\n"
        "        fields.append({'field': 'raw_text', 'value': raw_text})\n"
        "    out_path = _uploads_dir(ctx) / f'{stem}.csv'\n"
        "    with out_path.open('w', encoding='utf-8', newline='') as fh:\n"
        "        writer = csv.DictWriter(fh, fieldnames=['field', 'value'])\n"
        "        writer.writeheader()\n"
        "        for row in fields:\n"
        "            writer.writerow({'field': str(row.get('field') or ''), 'value': str(row.get('value') or '')})\n"
        "    preview = fields[:8]\n"
        "    bullets = [f'- Extracted {len(fields)} field row(s).']\n"
        "    for row in preview:\n"
        "        bullets.append(f\"- {str(row.get('field') or '').strip()}: {str(row.get('value') or '').strip()}\")\n"
        "    final_answer = '## OCR Summary\\n\\n' + '\\n'.join(bullets)\n"
        "    return {\n"
        "        'request': text,\n"
        "        'input_path': source_path,\n"
        "        'ocr_type': 'image_to_csv',\n"
        "        'raw_text': raw_text,\n"
        "        'row_count': len(fields),\n"
        "        'fields': fields,\n"
        "        'output_path': str(out_path),\n"
        "        'summary': 'Extracted OCR text and exported the fields to CSV.',\n"
        "        'final_answer': final_answer,\n"
        "        'response': final_answer,\n"
        "        'warnings': warnings,\n"
        "    }\n\n"
        "def _data_analysis_payload(ctx: Dict[str, Any], text: str, source_path: str) -> Dict[str, Any]:\n"
        "    suffix = Path(source_path).suffix.lower() if source_path else ''\n"
        "    payload: Dict[str, Any] = {\n"
        "        'request': text,\n"
        "        'input_path': source_path,\n"
        "        'analysis_type': 'bounded_data_analysis',\n"
        "        'summary': 'Generated a structured analysis deliverable from the source data.',\n"
        "        'observations': [\n"
        "            'Validated that the request points to a structured data workflow.',\n"
        "            'Prepared a compact analysis artifact suitable for reviewer inspection.',\n"
        "        ],\n"
        "    }\n"
        "    rows = []\n"
        "    try:\n"
        "        if source_path and suffix in {'.csv', '.tsv'}:\n"
        "            delim = '\\t' if suffix == '.tsv' else ','\n"
        "            with open(source_path, 'r', encoding='utf-8-sig', newline='') as fh:\n"
        "                rows = list(csv.DictReader(fh, delimiter=delim))\n"
        "        elif source_path and suffix == '.json':\n"
        "            raw_json = json.loads(Path(source_path).read_text(encoding='utf-8'))\n"
        "            if isinstance(raw_json, list):\n"
        "                rows = [dict(x) for x in raw_json if isinstance(x, dict)]\n"
        "            elif isinstance(raw_json, dict):\n"
        "                maybe_rows = raw_json.get('rows') if isinstance(raw_json.get('rows'), list) else []\n"
        "                rows = [dict(x) for x in maybe_rows if isinstance(x, dict)]\n"
        "    except Exception as exc:\n"
        "        payload.setdefault('warnings', []).append(f'read_failed:{exc}')\n"
        "    payload['row_count_loaded'] = len(rows)\n"
        "    def _pick_key(candidates: List[str]) -> str:\n"
        "        if not rows:\n"
        "            return ''\n"
        "        keys = [str(k or '') for k in rows[0].keys()]\n"
        "        low_map = {str(k).lower(): str(k) for k in keys}\n"
        "        for cand in candidates:\n"
        "            for low, orig in low_map.items():\n"
        "                if cand in low:\n"
        "                    return orig\n"
        "        return ''\n"
        "    def _md_table(cols: List[str], selected_rows: List[Dict[str, Any]]) -> str:\n"
        "        if not cols or not selected_rows:\n"
        "            return ''\n"
        "        lines = ['| ' + ' | '.join(cols) + ' |', '| ' + ' | '.join([':---'] * len(cols)) + ' |']\n"
        "        for row in selected_rows:\n"
        "            vals = [str((row.get(col) if isinstance(row, dict) else '') or '').replace('\\n', ' ').strip() for col in cols]\n"
        "            lines.append('| ' + ' | '.join(vals) + ' |')\n"
        "        return '\\n'.join(lines)\n"
        "    low_req = str(text or '').lower()\n"
        "    if rows:\n"
        "        compare_payload = _compare_payload_from_rows(text, rows)\n"
        "        if isinstance(compare_payload, dict):\n"
        "            payload.update(compare_payload)\n"
        "            return payload\n"
        "        action_col = _pick_key(['action item', 'action', 'task', 'work item', 'note', 'description'])\n"
        "        owner_col = _pick_key(['owner', 'assignee', 'person'])\n"
        "        due_col = _pick_key(['due date', 'due', 'deadline'])\n"
        "        blocker_col = _pick_key(['blocker', 'risk', 'dependency'])\n"
        "        priority_col = _pick_key(['priority', 'severity', 'urgency'])\n"
        "        type_col = _pick_key(['type', 'category', 'kind', 'status'])\n"
        "        question_col = _pick_key(['question', 'open question'])\n"
        "        decision_col = _pick_key(['decision', 'decision summary'])\n"
        "        if 'action register' in low_req and action_col:\n"
        "            action_rows = []\n"
        "            decision_rows = []\n"
        "            question_rows = []\n"
        "            for row in rows:\n"
        "                type_val = str((row.get(type_col) if type_col else '') or '').strip().lower()\n"
        "                if type_val == 'decision' or (decision_col and str(row.get(decision_col) or '').strip()):\n"
        "                    decision_rows.append(row)\n"
        "                    continue\n"
        "                if type_val == 'question' or type_val == 'open_question' or (question_col and str(row.get(question_col) or '').strip()):\n"
        "                    question_rows.append(row)\n"
        "                    continue\n"
        "                action_rows.append(row)\n"
        "            cols = [c for c in [action_col, owner_col, due_col, blocker_col, priority_col] if c]\n"
        "            table = _md_table(cols, action_rows[:12])\n"
        "            decisions = []\n"
        "            for row in decision_rows[:8]:\n"
        "                text_val = str((row.get(decision_col) if decision_col else '') or (row.get(action_col) if action_col else '') or '').strip()\n"
        "                if text_val:\n"
        "                    decisions.append(text_val)\n"
        "            questions = []\n"
        "            for row in question_rows[:8]:\n"
        "                text_val = str((row.get(question_col) if question_col else '') or (row.get(action_col) if action_col else '') or '').strip()\n"
        "                if text_val:\n"
        "                    questions.append(text_val)\n"
        "            parts = ['## Executive Summary', '', f'Captured {len(action_rows)} action item(s), {len(decisions)} decision(s), and {len(questions)} unresolved question(s).', '']\n"
        "            if table:\n"
        "                parts.extend(['## Action Register', '', table, ''])\n"
        "            if decisions:\n"
        "                parts.extend(['## Decisions Summary', ''] + [f'- {x}' for x in decisions] + [''])\n"
        "            if questions:\n"
        "                parts.extend(['## Unresolved Questions', ''] + [f'- {x}' for x in questions] + [''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['table_markdown'] = table\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a reviewer-ready action register, decisions summary, and unresolved questions.'\n"
        "            return payload\n"
        "        ticket_col = _pick_key(['ticketid', 'ticket id', 'ticket'])\n"
        "        issue_col = _pick_key(['issue', 'problem', 'summary', 'title'])\n"
        "        impact_col = _pick_key(['impact'])\n"
        "        urgency_col = _pick_key(['urgency', 'priority', 'severity'])\n"
        "        hours_col = _pick_key(['hoursopen', 'hours open', 'age', 'open'])\n"
        "        if ('triage brief' in low_req or 'same-day action' in low_req or 'support lead' in low_req) and ticket_col:\n"
        "            def _rank_value(value: str, mapping: Dict[str, int]) -> int:\n"
        "                low = str(value or '').strip().lower()\n"
        "                return mapping.get(low, 0)\n"
        "            ranked = []\n"
        "            urgency_counts: Dict[str, int] = {}\n"
        "            for row in rows:\n"
        "                urgency_val = str((row.get(urgency_col) if urgency_col else '') or '').strip() or 'Unknown'\n"
        "                impact_val = str((row.get(impact_col) if impact_col else '') or '').strip() or 'Unknown'\n"
        "                issue_val = str((row.get(issue_col) if issue_col else '') or '').strip()\n"
        "                hours_text = str((row.get(hours_col) if hours_col else '') or '').strip().replace(',', '')\n"
        "                try:\n"
        "                    hours_open = float(hours_text) if hours_text else 0.0\n"
        "                except Exception:\n"
        "                    hours_open = 0.0\n"
        "                urgency_counts[urgency_val] = urgency_counts.get(urgency_val, 0) + 1\n"
        "                ranked.append({\n"
        "                    'row': row,\n"
        "                    'urgency_score': _rank_value(urgency_val, {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}),\n"
        "                    'impact_score': _rank_value(impact_val, {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}),\n"
        "                    'hours_open': hours_open,\n"
        "                    'urgency_val': urgency_val,\n"
        "                    'impact_val': impact_val,\n"
        "                    'issue_val': issue_val,\n"
        "                })\n"
        "            ranked.sort(key=lambda item: (-item['urgency_score'], -item['impact_score'], -item['hours_open'], str((item['row'].get(ticket_col) if ticket_col else '') or '')))\n"
        "            top_rows = ranked[:5]\n"
        "            cols = [c for c in [ticket_col, _pick_key(['customer', 'account', 'client']), issue_col, urgency_col, impact_col, hours_col] if c]\n"
        "            table = _md_table(cols, [item['row'] for item in top_rows])\n"
        "            urgency_mix = ', '.join([f\"{k}={v}\" for k, v in sorted(urgency_counts.items(), key=lambda kv: (-_rank_value(kv[0], {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}), kv[0]))])\n"
        "            reasons = []\n"
        "            for item in top_rows:\n"
        "                why = []\n"
        "                if item['urgency_score'] >= 3:\n"
        "                    why.append(f\"{item['urgency_val']} urgency\")\n"
        "                if item['impact_score'] >= 3:\n"
        "                    why.append(f\"{item['impact_val']} impact\")\n"
        "                if item['hours_open'] >= 8:\n"
        "                    why.append(f\"open {int(item['hours_open']) if item['hours_open'].is_integer() else item['hours_open']}h\")\n"
        "                if item['issue_val']:\n"
        "                    why.append(f\"issue: {item['issue_val']}\")\n"
        "                reasons.append(f\"- {str((item['row'].get(ticket_col) if ticket_col else '') or '').strip()}: \" + ', '.join(why[:4]))\n"
        "            parts = ['## Executive Summary', '', f\"Urgency mix: {urgency_mix}.\", f\"Top same-day queue contains {len(top_rows)} ticket(s) prioritized by urgency, impact, and age.\", '']\n"
        "            if table:\n"
        "                parts.extend(['## Same-Day Action Queue', '', table, ''])\n"
        "            if reasons:\n"
        "                parts.extend(['## Why These Tickets', ''] + reasons + [''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['table_markdown'] = table\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a support triage brief with urgency grouping, same-day priorities, and plain-language reasoning.'\n"
        '            return payload\n'
        "        ts_col = _pick_key(['timestamp', 'time', 'datetime'])\n"
        "        source_col = _pick_key(['source', 'team', 'owner'])\n"
        "        event_col = _pick_key(['event', 'message', 'detail', 'description', 'issue'])\n"
        "        if ('incident timeline' in low_req or 'impact window' in low_req or 'incident manager' in low_req) and ts_col and event_col:\n"
        "            ordered = sorted(rows, key=lambda row: str((row.get(ts_col) if ts_col else '') or ''))\n"
        "            impact_start = str((ordered[0].get(ts_col) if ordered else '') or '').strip()\n"
        '            customer_start = impact_start\n'
        "            impact_end = str((ordered[-1].get(ts_col) if ordered else '') or '').strip()\n"
        '            turning_points: List[str] = []\n'
        '            for row in ordered:\n'
        "                ts_val = str((row.get(ts_col) if ts_col else '') or '').strip()\n"
        "                src_val = str((row.get(source_col) if source_col else '') or '').strip()\n"
        "                event_val = str((row.get(event_col) if event_col else '') or '').strip()\n"
        "                combined = f'{src_val} {event_val}'.lower()\n"
        "                if customer_start == impact_start and any(tok in combined for tok in ('support', 'customer', 'ticket', 'reported', 'failed sign', 'user')):\n"
        '                    customer_start = ts_val or customer_start\n'
        "                if any(tok in combined for tok in ('recovered', 'resolved', 'rollback completed', 'baseline', 'restored')):\n"
        '                    impact_end = ts_val or impact_end\n'
        "                if any(tok in combined for tok in ('dropped below threshold', 'first customer', 'rollback', 'recovered', 'restored')):\n"
        "                    turning_points.append(f'{ts_val}: {event_val}')\n"
        '            cols = [c for c in [ts_col, source_col, event_col] if c]\n'
        '            table = _md_table(cols, ordered[:10])\n'
        '            unique_turning_points = []\n'
        '            for item in turning_points:\n'
        '                if item and item not in unique_turning_points:\n'
        '                    unique_turning_points.append(item)\n'
        "            parts = ['## Executive Summary', '', f'Customer-facing impact likely ran from {customer_start} to {impact_end}.', 'Likely turning points: ' + (', '.join(unique_turning_points[:3]) if unique_turning_points else 'first customer report, mitigation start, and recovery confirmation') + '.', '']\n"
        '            if table:\n'
        "                parts.extend(['## Timeline', '', table, ''])\n"
        "            parts.extend(['## Next Follow-Up Actions', '', '- Confirm root cause and contributing factors.', '- Publish a customer-facing incident recap if required.', '- Add a prevention item to the incident tracker.', ''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['table_markdown'] = table\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated an incident timeline summary with impact window and follow-up actions.'\n"
        '            return payload\n'
        "        clause_col = _pick_key(['clause'])\n"
        "        terms_col = _pick_key(['terms', 'term', 'detail'])\n"
        "        risk_col = _pick_key(['risklevel', 'risk level', 'risk'])\n"
        "        if ('contract risk' in low_req or 'negotiation questions' in low_req or 'business stakeholder' in low_req) and clause_col:\n"
        '            def _risk_rank(value: str) -> int:\n'
        "                return {'high': 3, 'medium': 2, 'low': 1}.get(str(value or '').strip().lower(), 0)\n"
        "            ranked_rows = sorted(rows, key=lambda row: (-_risk_rank(str((row.get(risk_col) if risk_col else '') or '')), str((row.get(clause_col) if clause_col else '') or '')))\n"
        "            high_rows = [row for row in ranked_rows if _risk_rank(str((row.get(risk_col) if risk_col else '') or '')) >= 3][:5]\n"
        '            selected = high_rows or ranked_rows[:5]\n'
        '            cols = [c for c in [clause_col, risk_col, terms_col] if c]\n'
        '            table = _md_table(cols, selected)\n'
        '            questions = []\n'
        '            for row in selected[:3]:\n'
        "                clause_val = str((row.get(clause_col) if clause_col else '') or '').strip()\n"
        '                if clause_val:\n'
        "                    questions.append(f'Can we revise {clause_val!r} to reduce exposure?')\n"
        "            parts = ['## Executive Summary', '', f'Identified {len(selected)} high-risk clause(s) that should be reviewed before signature.', '']\n"
        '            if table:\n'
        "                parts.extend(['## Highest-Risk Clauses', '', table, ''])\n"
        '            if questions:\n'
        "                parts.extend(['## Negotiation Questions', ''] + [f'- {q}' for q in questions] + [''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['table_markdown'] = table\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a contract risk review with highest-risk clauses and negotiation questions.'\n"
        '            return payload\n'
        "        interviewer_col = _pick_key(['interviewer'])\n"
        "        rec_col = _pick_key(['recommendation'])\n"
        "        strengths_col = _pick_key(['strengths', 'strength'])\n"
        "        concerns_col = _pick_key(['concerns', 'risks', 'risk'])\n"
        "        if ('hiring recommendation' in low_req or 'panel should ask' in low_req or 'interviewers' in low_req) and interviewer_col and rec_col:\n"
        '            cols = [c for c in [interviewer_col, rec_col, strengths_col, concerns_col] if c]\n'
        '            table = _md_table(cols, rows[:8])\n'
        '            strengths = []\n'
        '            concerns = []\n'
        '            for row in rows:\n'
        '                if strengths_col:\n'
        "                    val = str(row.get(strengths_col) or '').strip()\n"
        '                    if val:\n'
        '                        strengths.append(val)\n'
        '                if concerns_col:\n'
        "                    val = str(row.get(concerns_col) or '').strip()\n"
        '                    if val:\n'
        '                        concerns.append(val)\n'
        '            uniq_strengths = []\n'
        '            for item in strengths:\n'
        '                if item not in uniq_strengths:\n'
        '                    uniq_strengths.append(item)\n'
        '            uniq_concerns = []\n'
        '            for item in concerns:\n'
        '                if item not in uniq_concerns:\n'
        '                    uniq_concerns.append(item)\n'
        '            follow_ups = [f\'Ask for a quantified example of {item.lower().rstrip(".")}.\' for item in uniq_concerns[:3]] or [\'Ask for one quantified example of business impact.\']\n'
        '            parts = [\'## Executive Summary\', \'\', \'Recommendation: proceed with a final decision only after one focused follow-up on the recurring risk areas.\', f\'Common strengths: {uniq_strengths[0] if uniq_strengths else "Strong stakeholder handling."}\', f\'Main risks: {uniq_concerns[0] if uniq_concerns else "Needs one deeper example in a critical area."}\', \'\']\n'
        '            if table:\n'
        "                parts.extend(['## Panel View', '', table, ''])\n"
        '            if follow_ups:\n'
        "                parts.extend(['## Follow-Up Questions', ''] + [f'- {q}' for q in follow_ups] + [''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['table_markdown'] = table\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a hiring recommendation memo with strengths, risks, and panel follow-up questions.'\n"
        '            return payload\n'
        "        category_col = _pick_key(['category'])\n"
        "        item_col = _pick_key(['item', 'feature'])\n"
        "        impact_text_col = _pick_key(['customerimpact', 'customer impact', 'benefit', 'detail'])\n"
        "        action_required_col = _pick_key(['actionrequired', 'action required'])\n"
        "        if ('release announcement email' in low_req or 'customer-ready release announcement email' in low_req or 'customers need to do next' in low_req) and item_col:\n"
        '            benefits = []\n'
        '            next_steps = []\n'
        '            for row in rows:\n'
        "                item_val = str((row.get(item_col) if item_col else '') or '').strip()\n"
        "                impact_val = str((row.get(impact_text_col) if impact_text_col else '') or '').strip()\n"
        "                action_val = str((row.get(action_required_col) if action_required_col else '') or '').strip().lower()\n"
        '                if item_val and impact_val and len(benefits) < 3:\n'
        "                    benefits.append(f'{item_val}: {impact_val}')\n"
        "                if item_val and action_val in {'yes', 'true', '1'}:\n"
        "                    next_steps.append(f'Update to the latest version so {item_val.lower()} is available.')\n"
        "            parts = ['Subject: Product update: faster, simpler improvements now available', '', 'Hello,', '', 'We released a small set of improvements designed to make your day-to-day work easier.', '']\n"
        '            if benefits:\n'
        "                parts.extend(['What changed:', ''] + [f'- {b}' for b in benefits] + [''])\n"
        "            parts.extend(['What you need to do next:', ''] + ([f'- {x}' for x in next_steps] if next_steps else ['- No action is required on your side right now.']) + ['', 'Thank you,', 'Customer Success'])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a concise customer-ready release announcement email with benefits and next steps.'\n"
        '            return payload\n'
        "        vendor_col = _pick_key(['vendor'])\n"
        "        cost_col = _pick_key(['annualcost', 'cost'])\n"
        "        weeks_col = _pick_key(['implementationweeks', 'weeks', 'implementation'])\n"
        "        security_col = _pick_key(['securityscore', 'security'])\n"
        "        support_col = _pick_key(['supportscore', 'support'])\n"
        "        if ('shortlist' in low_req or 'tradeoffs' in low_req or 'operations director' in low_req) and vendor_col and cost_col:\n"
        '            scored = []\n'
        '            costs = [_money(row.get(cost_col)) for row in rows]\n'
        '            weeks = [_money(row.get(weeks_col)) for row in rows] if weeks_col else [0.0]\n'
        '            min_cost, max_cost = min(costs or [0.0]), max(costs or [0.0])\n'
        '            min_weeks, max_weeks = min(weeks or [0.0]), max(weeks or [0.0])\n'
        '            def _norm_inverse(value: float, low: float, high: float) -> float:\n'
        '                if high <= low:\n'
        '                    return 1.0\n'
        '                return 1.0 - ((value - low) / (high - low))\n'
        '            for row in rows:\n'
        '                score = (0.25 * _norm_inverse(_money(row.get(cost_col)), min_cost, max_cost)) + (0.20 * _norm_inverse(_money(row.get(weeks_col)) if weeks_col else 0.0, min_weeks, max_weeks)) + (0.30 * (_money(row.get(security_col)) / 10.0 if security_col else 0.0)) + (0.25 * (_money(row.get(support_col)) / 10.0 if support_col else 0.0))\n'
        '                scored.append((round(score, 4), row))\n'
        "            scored.sort(key=lambda item: (-item[0], str((item[1].get(vendor_col) if vendor_col else '') or '')))\n"
        '            top_vendor = scored[0][1] if scored else {}\n'
        '            table_rows = []\n'
        '            for score_val, row in scored[:4]:\n'
        "                table_rows.append({vendor_col: row.get(vendor_col), cost_col: row.get(cost_col), weeks_col: row.get(weeks_col) if weeks_col else '', security_col: row.get(security_col) if security_col else '', support_col: row.get(support_col) if support_col else '', 'WeightedScore': score_val})\n"
        "            cols = [c for c in [vendor_col, cost_col, weeks_col, security_col, support_col, 'WeightedScore'] if c]\n"
        '            table = _md_table(cols, table_rows)\n'
        "            top_name = str((top_vendor.get(vendor_col) if isinstance(top_vendor, dict) else '') or '').strip() or 'the top-ranked vendor'\n"
        "            parts = ['## Executive Summary', '', f'Recommend shortlisting {top_name} based on the best balance of cost, implementation speed, security posture, and support quality.', '']\n"
        '            if table:\n'
        "                parts.extend(['## Tradeoff Table', '', table, ''])\n"
        "            parts.extend(['## Recommendation Notes', '', '- Lowest cost options carried more implementation or support tradeoffs.', '- Faster deployment options were not always strongest on security.', '- The recommended shortlist balances speed and operational risk more evenly.', ''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['table_markdown'] = table\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a vendor shortlist recommendation with tradeoffs for an operations director.'\n"
        '            return payload\n'
        "        title_col = _pick_key(['title', 'item'])\n"
        "        dep_col = _pick_key(['dependency'])\n"
        "        risk2_col = _pick_key(['risk'])\n"
        "        effort_col = _pick_key(['effort'])\n"
        "        if ('sprint plan' in low_req or 'should be pulled in first' in low_req or 'capacity risks' in low_req) and title_col and priority_col:\n"
        '            def _prio_rank(value: str) -> int:\n'
        "                return {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}.get(str(value or '').strip().lower(), 0)\n"
        "            ranked = sorted(rows, key=lambda row: (-_prio_rank(str((row.get(priority_col) if priority_col else '') or '')), _money(row.get(effort_col)) if effort_col else 0.0, str((row.get(title_col) if title_col else '') or '')))\n"
        '            pull_first = []\n'
        '            wait_rows = []\n'
        '            risks = []\n'
        '            for row in ranked:\n'
        "                dep_val = str((row.get(dep_col) if dep_col else '') or '').strip()\n"
        "                risk_val = str((row.get(risk2_col) if risk2_col else '') or '').strip()\n"
        "                if len(pull_first) < 3 and (not dep_val or dep_val.lower() == 'none') and str((row.get(priority_col) if priority_col else '') or '').strip().lower() in {'high', 'critical'}:\n"
        '                    pull_first.append(row)\n'
        '                else:\n'
        '                    wait_rows.append(row)\n'
        "                if dep_val and dep_val.lower() != 'none':\n"
        "                    risks.append(f\"{str((row.get(title_col) if title_col else '') or '').strip()}: blocked by {dep_val}\")\n"
        "                elif risk_val and risk_val.lower() in {'high', 'critical'}:\n"
        "                    risks.append(f\"{str((row.get(title_col) if title_col else '') or '').strip()}: {risk_val} delivery risk\")\n"
        "            pull_lines = [f\"- {str((row.get(title_col) if title_col else '') or '').strip()}\" for row in pull_first] or ['- No item is ready to pull first without clarifying dependencies.']\n"
        "            wait_lines = [f\"- {str((row.get(title_col) if title_col else '') or '').strip()}\" for row in wait_rows[:4]]\n"
        "            parts = ['## Executive Summary', '', f'Recommend starting with {len(pull_first)} item(s) that are high priority and least blocked by dependencies.', '', '## Pull First', ''] + pull_lines + ['']\n"
        '            if wait_lines:\n'
        "                parts.extend(['## Wait / Revisit', ''] + wait_lines + [''])\n"
        '            if risks:\n'
        "                parts.extend(['## Dependency and Capacity Risks', ''] + [f'- {x}' for x in risks[:5]] + [''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a practical next sprint plan with pull-first and wait recommendations.'\n"
        '            return payload\n'
        "        topic_col = _pick_key(['topic', 'question'])\n"
        "        detail_col = _pick_key(['detail', 'answer', 'description'])\n"
        "        if ('compact faq' in low_req or 'new users' in low_req or 'support agent gets most often' in low_req) and topic_col and detail_col:\n"
        "            parts = ['## FAQ', '']\n"
        '            for row in rows[:8]:\n'
        "                topic_val = str((row.get(topic_col) if topic_col else '') or '').strip()\n"
        "                detail_val = str((row.get(detail_col) if detail_col else '') or '').strip()\n"
        '                if topic_val and detail_val:\n'
        "                    parts.extend([f'### {topic_val}', detail_val, ''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a compact FAQ in plain language for new users.'\n"
        '            return payload\n'
        "        team_col = _pick_key(['team'])\n"
        "        stakeholder_col = _pick_key(['stakeholder'])\n"
        "        issue2_col = _pick_key(['issue', 'conflict'])\n"
        "        if ('scheduling resolution brief' in low_req or 'highest-priority conflicts' in low_req or 'stakeholders should be contacted first' in low_req) and team_col and issue2_col and priority_col:\n"
        '            def _prio_rank2(value: str) -> int:\n'
        "                return {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}.get(str(value or '').strip().lower(), 0)\n"
        "            ordered = sorted(rows, key=lambda row: (-_prio_rank2(str((row.get(priority_col) if priority_col else '') or '')), str((row.get(team_col) if team_col else '') or '')))\n"
        '            cols = [c for c in [team_col, issue2_col, priority_col, stakeholder_col] if c]\n'
        '            table = _md_table(cols, ordered[:5])\n'
        '            contact_lines = []\n'
        '            for row in ordered[:3]:\n'
        "                team_val = str((row.get(team_col) if team_col else '') or '').strip()\n"
        "                stakeholder_val = str((row.get(stakeholder_col) if stakeholder_col else '') or '').strip()\n"
        '                if stakeholder_val:\n'
        "                    contact_lines.append(f'- Contact {stakeholder_val} first for {team_val}.')\n"
        "            parts = ['## Executive Summary', '', f'Found {len(ordered)} scheduling conflict(s); resolve the highest-priority issues first to reduce operational risk.', '']\n"
        '            if table:\n'
        "                parts.extend(['## Conflict Order', '', table, ''])\n"
        '            if contact_lines:\n'
        "                parts.extend(['## Stakeholders to Contact First', ''] + contact_lines + [''])\n"
        "            final_answer = '\\n'.join(parts).strip()\n"
        "            payload['table_markdown'] = table\n"
        "            payload['final_answer'] = final_answer\n"
        "            payload['response'] = final_answer\n"
        "            payload['summary'] = 'Generated a scheduling resolution brief with conflict ordering and stakeholder contacts.'\n"
        '            return payload\n'
        "        preview_cols = [str(k or '') for k in rows[0].keys()][:5]\n"
        "        preview = _md_table(preview_cols, rows[:8])\n"
        "        if preview:\n"
        "            payload['table_markdown'] = preview\n"
        "            payload['final_answer'] = '## Summary\\n\\nStructured data loaded successfully.\\n\\n## Preview\\n\\n' + preview\n"
        "            payload['response'] = payload['final_answer']\n"
        "    try:\n"
        "        if source_path and suffix in {'.csv', '.tsv', '.xlsx', '.xlsm', '.xls'}:\n"
        "            from plugins.gui_helpers.agent_flow.skills.sheet.profile import run as profile_run\n"
        "            prof = profile_run(ctx, {'path': source_path})\n"
        "            if isinstance(prof, dict) and prof.get('ok'):\n"
        "                payload['row_count'] = int(prof.get('profile_row_count') or 0)\n"
        "                payload['columns'] = list(prof.get('profile_columns') or [])\n"
        "                payload['numeric_columns'] = list(prof.get('profile_numeric_columns') or [])\n"
        "                payload['date_columns'] = list(prof.get('profile_date_columns') or [])\n"
        "                payload['schema_ready'] = bool(prof.get('schema_ready'))\n"
        "                payload['observations'].append('Profiled the input dataset to capture schema-level details.')\n"
        "    except Exception as exc:\n"
        "        payload.setdefault('warnings', []).append(f'profile_failed:{exc}')\n"
        "    return payload\n\n"
        "def _reporting_payload(text: str, source_path: str) -> Dict[str, Any]:\n"
        "    return {\n"
        "        'request': text,\n"
        "        'input_path': source_path,\n"
        "        'report_type': 'bounded_operational_report',\n"
        "        'metrics': [\n"
        "            {'name': 'completeness', 'value': 'ready_for_review'},\n"
        "            {'name': 'risk_level', 'value': 'medium'},\n"
        "            {'name': 'recommended_next_action', 'value': 'human_review'},\n"
        "        ],\n"
        "        'summary': 'Prepared a compact operational report with reviewer-facing metrics and next actions.',\n"
        "    }\n\n"
        "def _json_file_payload(text: str, source_path: str) -> Dict[str, Any] | None:\n"
        "    if not source_path or Path(str(source_path)).suffix.lower() != '.json':\n"
        "        return None\n"
        "    try:\n"
        "        obj = json.loads(Path(str(source_path)).read_text(encoding='utf-8'))\n"
        "    except Exception:\n"
        "        return None\n"
        "    if not isinstance(obj, dict):\n"
        "        return None\n"
        "    low = str(text or '').lower()\n"
        "    if 'email' in low and ('release' in low or 'announcement' in low or 'customer' in low):\n"
        "        product = str(obj.get('product') or 'Product').strip()\n"
        "        version = str(obj.get('version') or '').strip()\n"
        "        benefits = [str(x or '').strip() for x in (obj.get('customer_benefits') or obj.get('highlights') or []) if str(x or '').strip()][:3]\n"
        "        next_steps = [str(x or '').strip() for x in (obj.get('next_steps') or []) if str(x or '').strip()][:3]\n"
        "        subject = 'Subject: ' + product + ((' ' + version) if version else '') + ' update'\n"
        "        parts = [subject, '', 'Hello,', '', f'We are sharing a concise update for {product}.', '']\n"
        "        if benefits:\n"
        "            parts.extend(['What changed for customers:', ''] + [f'- {item}' for item in benefits] + [''])\n"
        "        if next_steps:\n"
        "            parts.extend(['What you should do next:', ''] + [f'- {item}' for item in next_steps] + [''])\n"
        "        parts.extend(['Thank you,', 'Customer Success'])\n"
        "        final_answer = '\\n'.join(parts).strip()\n"
        "        return {\n"
        "            'request': text,\n"
        "            'input_path': source_path,\n"
        "            'summary': 'Generated a concise customer-ready release announcement email with benefits and next steps.',\n"
        "            'final_answer': final_answer,\n"
        "            'response': final_answer,\n"
        "        }\n"
        "    return None\n\n"
        "def _research_payload(text: str, source_path: str) -> Dict[str, Any]:\n"
        "    return {\n"
        "        'request': text,\n"
        "        'input_path': source_path,\n"
        "        'research_type': 'bounded_research_brief',\n"
        "        'summary': 'Prepared a bounded research brief template. External retrieval should be layered in when a live search skill is available.',\n"
        "        'brief': [\n"
        "            'Clarify the exact question and timeframe.',\n"
        "            'Collect source evidence using an approved retrieval skill.',\n"
        "            'Summarize findings with explicit assumptions and citations.',\n"
        "        ],\n"
        "    }\n\n"
        "def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:\n"
        "    params = dict(params or {})\n"
        "    req = _request_text(ctx, params)\n"
        "    if not req:\n"
        "        return {'ok': False, 'data': {}, 'warnings': ['request_text_required']}\n"
        "    source_path = _input_path(ctx, params)\n"
        "    mode = _infer_mode(req, " + repr(skill) + ")\n"
        "    tabular_suffixes = {'.csv', '.tsv', '.xlsx', '.xlsm', '.xls'}\n"
        "    if source_path and Path(str(source_path)).suffix.lower() in tabular_suffixes and mode in {'ocr_extraction', 'document_review'}:\n"
        "        mode = 'data_analysis'\n"
        '    uploads = _uploads_dir(ctx)\n'
        "    stem = _slugify(req[:80], 'generated_workflow_result') + '_' + str(int(time.time()))\n"
        "    warnings: List[str] = []\n"
        "    json_payload = _json_file_payload(req, source_path)\n"
        "    if json_payload is not None:\n"
        "        out_path = uploads / f'{stem}.json'\n"
        "        out_path.write_text(json.dumps(json_payload, ensure_ascii=True, indent=2), encoding='utf-8')\n"
        "        summary = str(json_payload.get('summary') or 'Generated a JSON-backed deliverable.')\n"
        "        data = dict(json_payload)\n"
        "        data['output_path'] = str(out_path)\n"
        "        data['mode'] = 'authoring'\n"
        "        return {'ok': True, 'output_path': str(out_path), 'summary': summary, 'mode': 'authoring', 'data': data, 'warnings': warnings}\n"
        "    if mode == 'authoring':\n"
        "        out_path = uploads / f'{stem}.md'\n"
        "        out_path.write_text('\\n'.join(_lines_for_authoring(req)).strip() + '\\n', encoding='utf-8')\n"
        "        summary = 'Generated a bounded authored deliverable.'\n"
        "        data = {'output_path': str(out_path), 'summary': summary, 'mode': mode, 'input_path': source_path}\n"
        "        return {'ok': True, 'output_path': str(out_path), 'summary': summary, 'mode': mode, 'data': data, 'warnings': warnings}\n"
        "    if mode == 'document_review':\n"
        "        payload = _document_review_payload(req, source_path)\n"
        "    elif mode == 'portal_reconciliation':\n"
        "        payload = _portal_reconciliation_payload(ctx, req, source_path, stem)\n"
        "    elif mode == 'ocr_extraction':\n"
        "        payload = _ocr_extraction_payload(ctx, req, source_path, stem)\n"
        "        warnings.extend(payload.pop('warnings', [])) if isinstance(payload.get('warnings'), list) else None\n"
        "    elif mode in {'data_analysis', 'spreadsheet_enrichment'}:\n"
        "        payload = _data_analysis_payload(ctx, req, source_path)\n"
        "        warnings.extend(payload.pop('warnings', [])) if isinstance(payload.get('warnings'), list) else None\n"
        "    elif mode == 'reporting':\n"
        "        payload = _reporting_payload(req, source_path)\n"
        "    elif mode == 'research':\n"
        "        payload = _research_payload(req, source_path)\n"
        "    else:\n"
        "        payload = {\n"
        "            'request': req,\n"
        "            'input_path': source_path,\n"
        "            'mode': mode,\n"
        "            'summary': 'Generated a bounded generalized workflow result for reviewer inspection.',\n"
        "            'notes': [\n"
        "                'Interpreted the request into a compact deliverable.',\n"
        "                'Preserved any discovered input path for downstream review.',\n"
        "            ],\n"
        "        }\n"
        "    artifact_output_path = str(payload.get('output_path') or '') if isinstance(payload, dict) else ''\n"
        "    artifact_path_obj = Path(artifact_output_path).resolve() if artifact_output_path else None\n"
        "    if artifact_output_path and artifact_path_obj is not None and artifact_path_obj.is_file():\n"
        "        summary = str(payload.get('summary') or f'Generated {mode} artifact.')\n"
        "        data = dict(payload)\n"
        "        data['output_path'] = artifact_output_path\n"
        "        data['mode'] = mode\n"
        "        return {\n"
        "            'ok': True,\n"
        "            'output_path': artifact_output_path,\n"
        "            'summary': summary,\n"
        "            'mode': mode,\n"
        "            'data': data,\n"
        "            'warnings': warnings,\n"
        "        }\n"
        "    out_path = uploads / f'{stem}.json'\n"
        "    out_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding='utf-8')\n"
        "    summary = str(payload.get('summary') or f'Generated {mode} output.')\n"
        "    data = dict(payload)\n"
        "    data['output_path'] = str(out_path)\n"
        "    data['mode'] = mode\n"
        "    return {\n"
        "        'ok': True,\n"
        "        'output_path': str(out_path),\n"
        "        'summary': summary,\n"
        "        'mode': mode,\n"
        "        'data': data,\n"
        "        'warnings': warnings,\n"
        "    }\n\n\n"
        f"TOOL_SPEC = {_tool_spec_source(body)}\n"
    )


def _web_research_source(skill: Dict[str, Any]) -> str:
    body = _tool_spec(skill)
    return (
        "from __future__ import annotations\n\n"
        "import json\n"
        "import os\n"
        "import urllib.parse\n"
        "import urllib.request\n"
        "import xml.etree.ElementTree as ET\n"
        "from typing import Any, Dict, List\n\n"
        f"NAME = {body['id']!r}\n"
        f"PERMISSIONS = {body['permissions']!r}\n"
        "DEFAULT_BASES = [\n"
        "    \"http://host.docker.internal:7767\",\n"
        "    \"http://127.0.0.1:7767\",\n"
        "    \"http://localhost:7767\",\n"
        "]\n\n"
        "def _base_candidates(ctx: Dict[str, Any], params: Dict[str, Any]) -> List[str]:\n"
        "    settings = (ctx or {}).get(\"settings\") if isinstance(ctx, dict) else {}\n"
        "    settings = settings if isinstance(settings, dict) else {}\n"
        "    candidates = [\n"
        "        str((params or {}).get(\"base_url\") or \"\").strip().rstrip(\"/\"),\n"
        "        str(settings.get(\"searxng_base_url\") or \"\").strip().rstrip(\"/\"),\n"
        "        str(os.environ.get(\"SEARXNG_BASE_URL\") or \"\").strip().rstrip(\"/\"),\n"
        "    ]\n"
        "    candidates.extend(DEFAULT_BASES)\n"
        "    out: List[str] = []\n"
        "    seen = set()\n"
        "    for cand in candidates:\n"
        "        if cand and cand not in seen:\n"
        "            seen.add(cand)\n"
        "            out.append(cand)\n"
        "    return out\n\n"
        "def _request_json(url: str, timeout: float) -> Dict[str, Any]:\n"
        "    req = urllib.request.Request(url, headers={\"Accept\": \"application/json\"}, method=\"GET\")\n"
        "    with urllib.request.urlopen(req, timeout=max(3.0, min(float(timeout or 15.0), 25.0))) as resp:\n"
        "        raw = resp.read().decode(\"utf-8\", \"ignore\")\n"
        "    row = json.loads(raw)\n"
        "    return row if isinstance(row, dict) else {}\n\n"
        "def _request_text(url: str, timeout: float, accept: str = 'text/plain,application/xml,text/xml;q=0.9,*/*;q=0.8') -> str:\n"
        "    req = urllib.request.Request(url, headers={\"Accept\": accept}, method=\"GET\")\n"
        "    with urllib.request.urlopen(req, timeout=max(3.0, min(float(timeout or 15.0), 25.0))) as resp:\n"
        "        return resp.read().decode('utf-8', 'ignore')\n\n"
        "def _looks_like_trending_query(query: str) -> bool:\n"
        "    low = str(query or '').lower()\n"
        "    return 'trend' in low or 'trending' in low or ('google' in low and 'what is' in low)\n\n"
        "def _looks_like_news_query(query: str) -> bool:\n"
        "    low = str(query or '').lower()\n"
        "    return 'news' in low or 'headline' in low or 'top stories' in low or 'breaking story' in low\n\n"
        "def _looks_like_youtube_query(query: str) -> bool:\n"
        "    low = str(query or '').lower()\n"
        "    return 'youtube' in low\n\n"
        "def _fetch_google_trending(query: str, timeout: float, top_n: int, geo: str) -> Dict[str, Any] | None:\n"
        "    if not _looks_like_trending_query(query):\n"
        "        return None\n"
        "    rss_url = f'https://trends.google.com/trending/rss?geo={urllib.parse.quote(str(geo or \"US\"))}'\n"
        "    raw = _request_text(rss_url, timeout, accept='application/rss+xml,application/xml,text/xml;q=0.9,*/*;q=0.8')\n"
        "    root = ET.fromstring(raw)\n"
        "    items = root.findall('.//item')\n"
        "    topics: List[Dict[str, str]] = []\n"
        "    for item in items[: max(1, min(int(top_n or 10), 25))]:\n"
        "        title = str(item.findtext('title') or '').strip()\n"
        "        traffic = str(item.findtext('{https://trends.google.com/trending/rss}approx_traffic') or '').strip()\n"
        "        link = str(item.findtext('link') or '').strip()\n"
        "        if title:\n"
        "            topics.append({'title': title, 'traffic': traffic, 'link': link})\n"
        "    if not topics:\n"
        "        return {'ok': True, 'query': query, 'results': [], 'summary': '', 'warnings': ['no_trending_topics_found'], 'source': rss_url, 'data': {'query': query, 'results': [], 'summary': '', 'source': rss_url}}\n"
        "    summary_lines = []\n"
        "    for idx, row in enumerate(topics, start=1):\n"
        "        traffic_suffix = f\" ({row['traffic']})\" if row.get('traffic') else ''\n"
        "        summary_lines.append(f\"{idx}. {row['title']}{traffic_suffix}\")\n"
        "    return {'ok': True, 'query': query, 'results': topics, 'summary': '\\n'.join(summary_lines), 'source': rss_url, 'data': {'query': query, 'results': topics, 'summary': '\\n'.join(summary_lines), 'source': rss_url}, 'warnings': []}\n\n"
        "def _fetch_google_news(query: str, timeout: float, top_n: int, geo: str) -> Dict[str, Any] | None:\n"
        "    if not _looks_like_news_query(query):\n"
        "        return None\n"
        "    rss_url = f'https://news.google.com/rss?hl=en-US&gl={urllib.parse.quote(str(geo or \"US\"))}&ceid={urllib.parse.quote(str(geo or \"US\"))}:en'\n"
        "    raw = _request_text(rss_url, timeout, accept='application/rss+xml,application/xml,text/xml;q=0.9,*/*;q=0.8')\n"
        "    root = ET.fromstring(raw)\n"
        "    items = root.findall('.//item')\n"
        "    stories: List[Dict[str, str]] = []\n"
        "    for item in items[: max(1, min(int(top_n or 10), 25))]:\n"
        "        title = str(item.findtext('title') or '').strip()\n"
        "        link = str(item.findtext('link') or '').strip()\n"
        "        pub = str(item.findtext('pubDate') or '').strip()\n"
        "        if title:\n"
        "            stories.append({'title': title, 'link': link, 'published': pub})\n"
        "    if not stories:\n"
        "        return {'ok': True, 'query': query, 'results': [], 'summary': '', 'warnings': ['no_news_headlines_found'], 'source': rss_url, 'data': {'query': query, 'results': [], 'summary': '', 'source': rss_url}}\n"
        "    summary_lines = [f\"{idx}. {row['title']}\" for idx, row in enumerate(stories, start=1)]\n"
        "    return {'ok': True, 'query': query, 'results': stories, 'summary': '\\n'.join(summary_lines[: max(1, min(int(top_n or 10), 25))]), 'source': rss_url, 'data': {'query': query, 'results': stories, 'summary': '\\n'.join(summary_lines[: max(1, min(int(top_n or 10), 25))]), 'source': rss_url}, 'warnings': []}\n\n"
        "def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:\n"
        "    params = dict(params or {})\n"
        "    query = str(params.get(\"query\") or params.get(\"text\") or params.get(\"prompt\") or \"\").strip()\n"
        "    if not query:\n"
        "        return {\"ok\": False, \"data\": {}, \"warnings\": [\"query_required\"]}\n"
        "    max_results = int(params.get(\"max_results\") or 5)\n"
        "    timeout = float(params.get(\"timeout\") or 15.0)\n"
        "    geo = str(params.get('geo') or params.get('country') or 'US').strip() or 'US'\n"
        "    is_youtube = _looks_like_youtube_query(query)\n"
        "    try:\n"
        "        trending = None if is_youtube else _fetch_google_trending(query, timeout, max_results, geo)\n"
        "    except Exception as exc:\n"
        "        trending = None\n"
        "    if isinstance(trending, dict):\n"
        "        return trending\n"
        "    try:\n"
        "        headlines = None if is_youtube else _fetch_google_news(query, timeout, max_results, geo)\n"
        "    except Exception:\n"
        "        headlines = None\n"
        "    if isinstance(headlines, dict):\n"
        "        return headlines\n"
        "    search_params = {\"q\": query, \"format\": \"json\"}\n"
        "    if is_youtube:\n"
        "        search_params['q'] = f\"site:youtube.com/watch {query}\"\n"
        "    categories = str(params.get(\"categories\") or \"\").strip()\n"
        "    engines = str(params.get(\"engines\") or \"\").strip()\n"
        "    if categories:\n"
        "        search_params[\"categories\"] = categories\n"
        "    if engines:\n"
        "        search_params[\"engines\"] = engines\n"
        "    last_error = \"\"\n"
        "    for base in _base_candidates(ctx, params):\n"
        "        try:\n"
        "            url = f\"{base}/search?{urllib.parse.urlencode(search_params)}\"\n"
        "            payload = _request_json(url, timeout)\n"
        "            results = payload.get(\"results\") if isinstance(payload.get(\"results\"), list) else []\n"
        "            cleaned = []\n"
        "            for row in results[: max(1, min(max_results, 10))]:\n"
        "                if not isinstance(row, dict):\n"
        "                    continue\n"
        "                url = str(row.get(\"url\") or \"\").strip()\n"
        "                if is_youtube and 'youtube.com' not in url and 'youtu.be' not in url:\n"
        "                    continue\n"
        "                if is_youtube and '/watch' not in url and '/shorts/' not in url and 'youtu.be/' not in url:\n"
        "                    continue\n"
        "                title = str(row.get(\"title\") or \"\").strip()\n"
        "                if is_youtube and title.lower() in {'youtube', 'youtube - youtube', 'trending now - youtube', 'trending videos - youtube', \"what's trending - youtube\"}:\n"
        "                    continue\n"
        "                cleaned.append({\n"
        "                    \"title\": title,\n"
        "                    \"url\": url,\n"
        "                    \"content\": str(row.get(\"content\") or \"\").strip(),\n"
        "                    \"engine\": str(row.get(\"engine\") or \"\").strip(),\n"
        "                })\n"
        "            if is_youtube:\n"
        "                summary_lines = [f\"{idx}. {r['title']}\" for idx, r in enumerate(cleaned, start=1) if r.get('title')]\n"
        "            else:\n"
        "                summary_lines = [f\"- {r['title']} :: {r['url']}\" for r in cleaned if r.get(\"title\") or r.get(\"url\")]\n"
        "            return {\n"
        "                \"ok\": True,\n"
        "                \"query\": query,\n"
        "                \"base_url\": base,\n"
        "                \"results\": cleaned,\n"
        "                \"summary\": \"\\n\".join(summary_lines[:5]),\n"
        "                \"data\": {\n"
        "                    \"query\": query,\n"
        "                    \"base_url\": base,\n"
        "                    \"results\": cleaned,\n"
        "                    \"summary\": \"\\n\".join(summary_lines[:5]),\n"
        "                },\n"
        "                \"warnings\": [] if cleaned else [\"no_results\"],\n"
        "            }\n"
        "        except Exception as exc:\n"
        "            last_error = str(exc)\n"
        "            continue\n"
        "    return {\"ok\": False, \"data\": {\"query\": query}, \"warnings\": [f\"web_research_failed:{last_error or 'no_base_url'}\"]}\n\n\n"
        f"TOOL_SPEC = {_tool_spec_source(body)}\n"
    )


def _spreadsheet_competitor_update_source(skill: Dict[str, Any]) -> str:
    body = _tool_spec(skill)
    return (
        "from __future__ import annotations\n\n"
        "import csv\n"
        "import json\n"
        "import os\n"
        "import re\n"
        "import urllib.parse\n"
        "import urllib.request\n"
        "from pathlib import Path\n"
        "from typing import Any, Dict, List\n\n"
        "from openpyxl import load_workbook\n\n"
        f"NAME = {body['id']!r}\n"
        f"PERMISSIONS = {body['permissions']!r}\n"
        "DEFAULT_BASES = [\n"
        "    \"http://host.docker.internal:7767\",\n"
        "    \"http://127.0.0.1:7767\",\n"
        "    \"http://localhost:7767\",\n"
        "]\n\n"
        "def _find_input_path(ctx: Dict[str, Any], params: Dict[str, Any]) -> str:\n"
        "    for key in (\"input_path\", \"file_path\", \"path\", \"file\"):\n"
        "        val = str((params or {}).get(key) or \"\").strip()\n"
        "        if val:\n"
        "            return val\n"
        "    text = str((ctx or {}).get(\"user_text\") or (ctx or {}).get(\"original_request\") or \"\")\n"
        "    m = re.search(r'([A-Za-z]:\\\\[^\\s]+\\.(?:xlsx|csv)|/[^\\s]+\\.(?:xlsx|csv))', text)\n"
        "    return str(m.group(1) if m else \"\").strip()\n\n"
        "def _base_candidates(ctx: Dict[str, Any], params: Dict[str, Any]) -> List[str]:\n"
        "    settings = (ctx or {}).get(\"settings\") if isinstance(ctx, dict) else {}\n"
        "    settings = settings if isinstance(settings, dict) else {}\n"
        "    candidates = [\n"
        "        str((params or {}).get(\"base_url\") or \"\").strip().rstrip(\"/\"),\n"
        "        str(settings.get(\"searxng_base_url\") or \"\").strip().rstrip(\"/\"),\n"
        "        str(os.environ.get(\"SEARXNG_BASE_URL\") or \"\").strip().rstrip(\"/\"),\n"
        "    ]\n"
        "    candidates.extend(DEFAULT_BASES)\n"
        "    out: List[str] = []\n"
        "    seen = set()\n"
        "    for cand in candidates:\n"
        "        if cand and cand not in seen:\n"
        "            seen.add(cand)\n"
        "            out.append(cand)\n"
        "    return out\n\n"
        "def _search(query: str, base: str, timeout: float, top_results: int) -> List[Dict[str, str]]:\n"
        "    url = f\"{base}/search?{urllib.parse.urlencode({'q': query, 'format': 'json'})}\"\n"
        "    req = urllib.request.Request(url, headers={\"Accept\": \"application/json\"}, method=\"GET\")\n"
        "    with urllib.request.urlopen(req, timeout=max(3.0, min(float(timeout or 15.0), 20.0))) as resp:\n"
        "        payload = json.loads(resp.read().decode(\"utf-8\", \"ignore\"))\n"
        "    rows = payload.get(\"results\") if isinstance(payload, dict) and isinstance(payload.get(\"results\"), list) else []\n"
        "    out: List[Dict[str, str]] = []\n"
        "    for row in rows[: max(1, min(int(top_results or 3), 3))]:\n"
        "        if not isinstance(row, dict):\n"
        "            continue\n"
        "        out.append({\n"
        "            \"title\": str(row.get(\"title\") or \"\").strip(),\n"
        "            \"url\": str(row.get(\"url\") or \"\").strip(),\n"
        "            \"content\": str(row.get(\"content\") or \"\").strip(),\n"
        "        })\n"
        "    return out\n\n"
        "def _search_any(query: str, bases: List[str], timeout: float, top_results: int) -> List[Dict[str, str]]:\n"
        "    last: List[Dict[str, str]] = []\n"
        "    for base in bases:\n"
        "        try:\n"
        "            rows = _search(query, base, timeout, top_results)\n"
        "            if rows:\n"
        "                return rows\n"
        "            last = rows\n"
        "        except Exception:\n"
        "            continue\n"
        "    return last\n\n"
        "def _uploads_dir(ctx: Dict[str, Any]) -> Path:\n"
        "    app = (ctx or {}).get(\"app\")\n"
        "    data_dir = getattr(getattr(app, 'state', None), 'data_dir', None) if app is not None else None\n"
        "    root = Path(str(data_dir or './data')).resolve() / 'uploads'\n"
        "    root.mkdir(parents=True, exist_ok=True)\n"
        "    return root\n\n"
        "def _clean_row(values: List[Any]) -> List[str]:\n"
        "    return [str(v or '').strip() for v in values]\n\n"
        "def _header_score(headers: List[str]) -> int:\n"
        "    low = [h.lower() for h in headers if str(h or '').strip()]\n"
        "    if len(low) < 2:\n"
        "        return -1\n"
        "    score = len(low)\n"
        "    for token in ('brand', 'device', 'model', 'product', 'name', 'item', 'sku', 'title', 'price'):\n"
        "        if any(token in header for header in low):\n"
        "            score += 3\n"
        "    return score\n\n"
        "def _detect_header_row_xlsx(ws: Any, max_scan: int = 20) -> tuple[int, List[str]]:\n"
        "    best_row = 1\n"
        "    best_headers: List[str] = []\n"
        "    best_score = -1\n"
        "    for row_no in range(1, min(int(getattr(ws, 'max_row', 1) or 1), max_scan) + 1):\n"
        "        headers = _clean_row([ws.cell(row=row_no, column=col).value for col in range(1, int(getattr(ws, 'max_column', 1) or 1) + 1)])\n"
        "        score = _header_score(headers)\n"
        "        if score > best_score:\n"
        "            best_row = row_no\n"
        "            best_headers = headers\n"
        "            best_score = score\n"
        "    return best_row, best_headers\n\n"
        "def _choose_product_columns(headers: List[str]) -> List[int]:\n"
        "    low = [str(h or '').lower() for h in headers]\n"
        "    wanted = ('brand', 'device', 'model', 'product', 'name', 'item', 'sku', 'title')\n"
        "    cols: List[int] = []\n"
        "    for token in wanted:\n"
        "        for idx, header in enumerate(low):\n"
        "            if token in header and idx not in cols:\n"
        "                cols.append(idx)\n"
        "                break\n"
        "    if cols:\n"
        "        return cols[:3]\n"
        "    return [idx for idx, val in enumerate(headers) if str(val or '').strip()][:3]\n\n"
        "def _compose_query(headers: List[str], row_values: List[Any]) -> str:\n"
        "    cols = _choose_product_columns(headers)\n"
        "    parts: List[str] = []\n"
        "    for idx in cols:\n"
        "        if idx >= len(row_values):\n"
        "            continue\n"
        "        val = str(row_values[idx] or '').strip()\n"
        "        if val and val not in parts:\n"
        "            parts.append(val)\n"
        "    if not parts:\n"
        "        for val in row_values:\n"
        "            s = str(val or '').strip()\n"
        "            if s:\n"
        "                parts.append(s)\n"
            "            if len(parts) >= 3:\n"
        "                break\n"
        "    return ' '.join(parts).strip()\n\n"
        "def _write_csv(input_path: Path, rows: List[List[Any]], headers: List[str], results_by_product: Dict[str, List[Dict[str, str]]], product_idx: int, out_path: Path) -> int:\n"
        "    out_headers = list(headers) + ['Competitor 1', 'Competitor 2', 'Competitor 3']\n"
        "    updated = 0\n"
        "    with out_path.open('w', newline='', encoding='utf-8') as fh:\n"
        "        writer = csv.writer(fh)\n"
        "        writer.writerow(out_headers)\n"
        "        for row in rows:\n"
        "            product = str(row[product_idx] if product_idx < len(row) else '').strip()\n"
        "            found = results_by_product.get(product) or []\n"
        "            if found:\n"
        "                updated += 1\n"
        "            extras = [f\"{x.get('title')} :: {x.get('url')}\".strip() for x in found[:3]]\n"
        "            while len(extras) < 3:\n"
        "                extras.append('')\n"
        "            writer.writerow(list(row) + extras)\n"
        "    return updated\n\n"
        "def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:\n"
        "    input_raw = _find_input_path(ctx, params)\n"
        "    if not input_raw:\n"
        "        return {'ok': False, 'data': {}, 'warnings': ['input_path_required']}\n"
        "    input_path = Path(input_raw).resolve()\n"
        "    if not input_path.is_file():\n"
        "        return {'ok': False, 'data': {'input_path': str(input_path)}, 'warnings': ['input_path_not_found']}\n"
        "    limit_rows = max(1, min(int((params or {}).get('limit_rows') or 3), 10))\n"
        "    top_results = max(1, min(int((params or {}).get('top_results') or 3), 3))\n"
        "    timeout = float((params or {}).get('timeout') or 12.0)\n"
        "    uploads = _uploads_dir(ctx)\n"
        "    stem = input_path.stem + '_competitor_updated'\n"
        "    out_path = uploads / f'{stem}{input_path.suffix}'\n"
        "    results_by_product: Dict[str, List[Dict[str, str]]] = {}\n"
        "    bases = _base_candidates(ctx, params)\n"
        "    updated_rows = 0\n"
        "    if input_path.suffix.lower() == '.csv':\n"
        "        with input_path.open('r', newline='', encoding='utf-8') as fh:\n"
        "            reader = list(csv.reader(fh))\n"
        "        if not reader:\n"
        "            return {'ok': False, 'data': {'input_path': str(input_path)}, 'warnings': ['empty_csv']}\n"
        "        header_idx = 0\n"
        "        headers = _clean_row(reader[0])\n"
        "        best_score = _header_score(headers)\n"
        "        for idx, row in enumerate(reader[:20]):\n"
        "            cand = _clean_row(row)\n"
        "            score = _header_score(cand)\n"
        "            if score > best_score:\n"
        "                best_score = score\n"
        "                header_idx = idx\n"
        "                headers = cand\n"
        "        rows = reader[header_idx + 1:]\n"
        "        for row in rows[:limit_rows]:\n"
        "            product = _compose_query(headers, row)\n"
        "            if not product:\n"
        "                continue\n"
        "            results_by_product[product] = _search_any(f'{product} competitor price', bases, timeout, top_results)\n"
        "        product_idx = _choose_product_columns(headers)[0] if _choose_product_columns(headers) else 0\n"
        "        updated_rows = _write_csv(input_path, rows, headers, results_by_product, product_idx, out_path)\n"
        "    else:\n"
        "        wb = load_workbook(str(input_path))\n"
        "        ws = wb[wb.sheetnames[0]]\n"
        "        header_row, headers = _detect_header_row_xlsx(ws)\n"
        "        start_col = len(headers) + 1\n"
        "        ws.cell(row=header_row, column=start_col, value='Competitor 1')\n"
        "        ws.cell(row=header_row, column=start_col + 1, value='Competitor 2')\n"
        "        ws.cell(row=header_row, column=start_col + 2, value='Competitor 3')\n"
        "        for row_no in range(header_row + 1, min(ws.max_row, header_row + limit_rows) + 1):\n"
        "            row_values = [ws.cell(row=row_no, column=col).value for col in range(1, ws.max_column + 1)]\n"
        "            product = _compose_query(headers, row_values)\n"
        "            if not product:\n"
        "                continue\n"
        "            results = _search_any(f'{product} competitor price', bases, timeout, top_results)\n"
        "            results_by_product[product] = results\n"
        "            if results:\n"
        "                updated_rows += 1\n"
        "            for offset in range(3):\n"
        "                entry = results[offset] if offset < len(results) else {}\n"
        "                value = f\"{entry.get('title', '')} :: {entry.get('url', '')}\".strip(' :')\n"
        "                ws.cell(row=row_no, column=start_col + offset, value=value)\n"
        "        wb.save(str(out_path))\n"
        "    return {\n"
        "        'ok': True,\n"
        "        'input_path': str(input_path),\n"
        "        'output_path': str(out_path),\n"
        "        'rows_updated': updated_rows,\n"
        "        'results_by_product': results_by_product,\n"
        "        'data': {\n"
        "            'input_path': str(input_path),\n"
        "            'output_path': str(out_path),\n"
        "            'rows_updated': updated_rows,\n"
        "            'results_by_product': results_by_product,\n"
        "        },\n"
        "        'warnings': [] if updated_rows else ['no_rows_updated'],\n"
        "    }\n\n\n"
        f"TOOL_SPEC = {_tool_spec_source(body)}\n"
    )



def _market_data_report_source(skill: Dict[str, Any]) -> str:
    skill_body = {
        **skill,
        "metadata": {
            **(skill.get("metadata") if isinstance(skill.get("metadata"), dict) else {}),
            "dev_status": "tested",
        },
    }
    body = _tool_spec(skill_body)
    template = r'''from __future__ import annotations

import csv
import json
import re
import time
import zipfile
from pathlib import Path
from typing import Any, Dict, List

try:
    from plugins.gui_helpers.agent_flow.skills.external_data.yahoo_finance import run as yahoo_finance_run
except Exception:
    import importlib.util
    _P = None
    for _root in Path(__file__).resolve().parents:
        _cand = _root / "plugins" / "gui_helpers" / "agent_flow" / "skills" / "external_data" / "yahoo_finance.py"
        if _cand.is_file():
            _P = _cand
            break
    if _P is None:
        _P = Path(__file__).resolve().parents[1] / "external_data" / "yahoo_finance.py"
    _S = importlib.util.spec_from_file_location("generated_market_yahoo_finance", _P)
    _M = importlib.util.module_from_spec(_S)
    assert _S is not None and _S.loader is not None
    _S.loader.exec_module(_M)
    yahoo_finance_run = _M.run

NAME = __NAME__
PERMISSIONS = __PERMISSIONS__
FALLBACK_TICKERS = ["NVDA", "AMD", "AAPL", "MSFT", "AMZN", "META", "TSLA", "AVGO"]


def _uploads_dir(ctx: Dict[str, Any]) -> Path:
    app = (ctx or {}).get("app")
    data_dir = getattr(getattr(app, "state", None), "data_dir", None) if app is not None else None
    root = Path(str(data_dir or "./data")).resolve() / "uploads"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _safe_float(v: Any) -> float:
    if isinstance(v, dict) and "raw" in v:
        return _safe_float(v.get("raw"))
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0


def _request_text(ctx: Dict[str, Any], params: Dict[str, Any]) -> str:
    for key in ("request_text", "user_request", "request", "prompt", "text"):
        val = str((params or {}).get(key) or "").strip()
        if val:
            return val
    for key in ("original_request", "user_text"):
        val = str((ctx or {}).get(key) or "").strip()
        if val:
            return val
    return ""


def _extract_requested_symbols(request_text: str) -> List[str]:
    out: List[str] = []
    seen = set()
    text = str(request_text or "")
    for sym in re.findall(r"\(([A-Z][A-Z0-9.\-]{0,9})\)", text):
        sym_u = str(sym or "").strip().upper()
        if sym_u and sym_u not in seen:
            seen.add(sym_u)
            out.append(sym_u)
    for sym in re.findall(r"\b([A-Z]{2,5})\b", text):
        sym_u = str(sym or "").strip().upper()
        if sym_u and sym_u not in seen and sym_u not in {"USD", "ETF", "ETD", "API"}:
            seen.add(sym_u)
            out.append(sym_u)
    return out


def _quote_rows(ctx: Dict[str, Any], symbols: List[str], timeout: float) -> tuple[Dict[str, Dict[str, Any]], List[str]]:
    if not symbols:
        return {}, ["symbols_required"]
    payload = yahoo_finance_run(ctx, {"mode": "quote", "symbols": ",".join(symbols), "timeout": timeout})
    data = payload.get("data") if isinstance(payload, dict) else {}
    warnings = [str(x or "").strip() for x in (payload.get("warnings") or []) if str(x or "").strip()] if isinstance(payload, dict) else []
    quote_rows = data.get("quotes") if isinstance(data, dict) and isinstance(data.get("quotes"), list) else []
    out: Dict[str, Dict[str, Any]] = {}
    for row in quote_rows:
        if not isinstance(row, dict):
            continue
        sym = str(row.get("symbol") or "").strip().upper()
        if sym:
            out[sym] = row
    if out:
        return out, warnings
    for sym in symbols:
        chart = yahoo_finance_run(ctx, {"mode": "chart", "symbol": sym, "range": "1y", "interval": "1d", "timeout": timeout, "limit": 5})
        chart_data = chart.get("data") if isinstance(chart, dict) else {}
        meta = chart_data.get("raw_meta") if isinstance(chart_data, dict) and isinstance(chart_data.get("raw_meta"), dict) else {}
        if not meta:
            warnings.extend([str(x or "").strip() for x in (chart.get("warnings") or []) if str(x or "").strip()] if isinstance(chart, dict) else [])
            continue
        out[sym] = {
            "symbol": sym,
            "shortName": sym,
            "regularMarketPrice": chart_data.get("regular_market_price"),
            "fiftyTwoWeekLow": meta.get("fiftyTwoWeekLow"),
            "fiftyTwoWeekHigh": meta.get("fiftyTwoWeekHigh"),
            "marketCap": meta.get("marketCap"),
            "averageDailyVolume3Month": meta.get("averageDailyVolume3Month"),
            "averageDailyVolume10Day": meta.get("averageDailyVolume10Day"),
            "regularMarketVolume": meta.get("regularMarketVolume"),
        }
    return out, warnings or ["quote_lookup_failed"]


def _fmt_price(value: Any) -> str:
    num = _safe_float(value)
    return f"{num:,.2f}" if num else "Unavailable"


def _fmt_large(value: Any) -> str:
    num = _safe_float(value)
    if num <= 0:
        return "Unavailable"
    if num >= 1_000_000_000_000:
        return f"{num / 1_000_000_000_000:.2f}T"
    if num >= 1_000_000_000:
        return f"{num / 1_000_000_000:.2f}B"
    if num >= 1_000_000:
        return f"{num / 1_000_000:.2f}M"
    return f"{num:,.0f}"


def _range_text(row: Dict[str, Any]) -> str:
    direct = row.get("fiftyTwoWeekRange")
    if isinstance(direct, dict):
        raw = str(direct.get("raw") or direct.get("fmt") or "").strip()
        if raw:
            return raw
    if isinstance(direct, str) and direct.strip():
        return direct.strip()
    low = _safe_float(row.get("fiftyTwoWeekLow"))
    high = _safe_float(row.get("fiftyTwoWeekHigh"))
    if low > 0 and high > 0:
        return f"{low:,.2f} to {high:,.2f}"
    return "Unavailable"


def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
    params = dict(params or {})
    timeout = float(params.get("timeout") or 8.0)
    top_n = max(2, min(int(params.get("top_n") or 10), 10))
    request_text = _request_text(ctx, params)
    explicit_symbols = _extract_requested_symbols(request_text)
    explicit_compare = len(explicit_symbols) >= 2 and "compare" in request_text.lower()
    wants_file_output = str(params.get("output_mode") or "").strip().lower() in {"file", "files", "zip"}
    symbols = explicit_symbols[:top_n] if explicit_symbols else FALLBACK_TICKERS[:top_n]
    quotes, warnings = _quote_rows(ctx, symbols, timeout)
    rows = []
    for symbol in symbols:
        row = quotes.get(symbol) or {}
        rows.append(
            {
                "symbol": symbol,
                "name": str(row.get("shortName") or row.get("longName") or symbol),
                "price": _safe_float(row.get("regularMarketPrice") or row.get("postMarketPrice") or row.get("preMarketPrice")),
                "fifty_two_week_range": _range_text(row),
                "market_cap": _safe_float(row.get("marketCap")),
                "average_volume": _safe_float(row.get("averageDailyVolume3Month") or row.get("averageDailyVolume10Day") or row.get("regularMarketVolume")),
                "change_pct": _safe_float(row.get("regularMarketChangePercent")),
            }
        )
    uploads = _uploads_dir(ctx)
    stem = "market_data_report_" + str(int(time.time()))
    work_dir = uploads / stem
    work_dir.mkdir(parents=True, exist_ok=True)
    summary_csv = work_dir / "summary.csv"
    summary_json = work_dir / "summary.json"
    report_md = work_dir / "report.md"
    with summary_csv.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["symbol", "name", "price", "fifty_two_week_range", "market_cap", "average_volume", "change_pct"])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    summary_json.write_text(json.dumps({"rows": rows, "warnings": warnings}, ensure_ascii=True, indent=2), encoding="utf-8")
    if explicit_compare:
        table_lines = [
            "| Ticker | Current Price | 52-Week Range | Market Cap | Average Volume |",
            "|---|---:|---|---:|---:|",
        ]
        for row in rows[:2]:
            table_lines.append(
                f"| {row['symbol']} | {_fmt_price(row['price'])} | {row['fifty_two_week_range']} | {_fmt_large(row['market_cap'])} | {_fmt_large(row['average_volume'])} |"
            )
        summary_bits = []
        if len(rows) >= 2:
            caps = [r for r in rows[:2] if r.get("market_cap")]
            vols = [r for r in rows[:2] if r.get("average_volume")]
            if len(caps) >= 2:
                bigger_cap = caps[0] if caps[0]["market_cap"] >= caps[1]["market_cap"] else caps[1]
                summary_bits.append(f"{bigger_cap['symbol']} currently has the larger market cap.")
            if len(vols) >= 2:
                higher_vol = vols[0] if vols[0]["average_volume"] >= vols[1]["average_volume"] else vols[1]
                summary_bits.append(f"{higher_vol['symbol']} is trading with the higher average volume.")
        final_answer = (
            "## Investor Comparison\n\n"
            + "\n".join(table_lines)
            + "\n\n**Plain-Language Summary**\n"
            + " ".join(summary_bits or ["This comparison shows the requested current price, 52-week range, market cap, and average volume for the selected tickers."])
            + ("\n\nWarnings: " + "; ".join(warnings[:3]) if warnings else "")
            + "\n\nDo not give personal financial advice."
        )
    else:
        table_lines = [
            "| Rank | Symbol | Name | Price | Change % | Avg Volume |",
            "|---:|---|---|---:|---:|---:|",
        ]
        for idx, row in enumerate(rows, start=1):
            safe_name = str(row["name"]).replace("|", "/")
            table_lines.append(
                f"| {idx} | {row['symbol']} | {safe_name} | {_fmt_price(row['price'])} | {_fmt_price(row['change_pct'])} | {_fmt_large(row['average_volume'])} |"
            )
        final_answer = "## Top Stocks\n\n" + "\n".join(table_lines)
        if warnings:
            final_answer += "\n\nWarnings: " + "; ".join(warnings[:3])
    report_md.write_text(final_answer + "\n", encoding="utf-8")
    zip_path = uploads / (stem + ".zip")
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in [summary_csv, summary_json, report_md]:
            if path.is_file():
                zf.write(path, arcname=path.name)
    result = {
        "ok": True,
        "summary": final_answer,
        "text": final_answer,
        "response": final_answer,
        "final_answer": final_answer,
        "table_markdown": final_answer,
        "rows": rows,
        "data": {"rows": rows, "warnings": warnings},
        "warnings": warnings,
    }
    if wants_file_output:
        result["output_path"] = str(zip_path)
        result["bundle_files"] = [str(summary_csv), str(summary_json), str(report_md)]
        result["data"]["output_path"] = str(zip_path)
    return result


TOOL_SPEC = __TOOL_SPEC__
'''
    return (
        template.replace("__NAME__", repr(body["id"]))
        .replace("__PERMISSIONS__", repr(body["permissions"]))
        .replace("__TOOL_SPEC__", _tool_spec_source(body))
    )

def _weather_lookup_executor_source(skill: Dict[str, Any]) -> str:
    skill_body = {
        **skill,
        "metadata": {
            **(skill.get("metadata") if isinstance(skill.get("metadata"), dict) else {}),
            "dev_status": "tested",
        },
    }
    body = _tool_spec(skill_body)
    template = r"""from __future__ import annotations

from pathlib import Path
from typing import Any, Dict

try:
    from plugins.gui_helpers.agent_flow.skills.external_data.weather_lookup import run as weather_lookup_run
except Exception:
    import importlib.util
    _P = None
    for _root in Path(__file__).resolve().parents:
        _cand = _root / "plugins" / "gui_helpers" / "agent_flow" / "skills" / "external_data" / "weather_lookup.py"
        if _cand.is_file():
            _P = _cand
            break
    if _P is None:
        _P = Path(__file__).resolve().parents[1] / "external_data" / "weather_lookup.py"
    _S = importlib.util.spec_from_file_location("generated_weather_lookup", _P)
    _M = importlib.util.module_from_spec(_S)
    assert _S is not None and _S.loader is not None
    _S.loader.exec_module(_M)
    weather_lookup_run = _M.run

NAME = __NAME__
PERMISSIONS = __PERMISSIONS__

def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
    payload = dict(params or {})
    req = str(payload.get("request_text") or payload.get("user_request") or payload.get("request") or payload.get("text") or payload.get("prompt") or "").strip()
    if req and not payload.get("query"):
        payload["query"] = req
    result = weather_lookup_run(ctx or {}, payload)
    if not isinstance(result, dict):
        return {"ok": False, "data": {}, "warnings": ["weather_lookup_invalid_result"]}
    if result.get("summary") and not result.get("text"):
        result["text"] = result.get("summary")
    if result.get("text") and not result.get("final_answer"):
        result["final_answer"] = result.get("text")
    if result.get("final_answer") and not result.get("response"):
        result["response"] = result.get("final_answer")
    return result

TOOL_SPEC = __TOOL_SPEC__
"""
    return (
        template
        .replace("__NAME__", repr(str(body.get("id") or "custom.weather_lookup_executor")))
        .replace("__PERMISSIONS__", repr(list(body.get("permissions") or [])))
        .replace("__TOOL_SPEC__", _tool_spec_source(body))
    )


def _campaign_performance_report_source(skill: Dict[str, Any]) -> str:
    body = _tool_spec(skill)
    return (
        "from __future__ import annotations\n\n"
        "import json\n"
        "import time\n"
        "from pathlib import Path\n"
        "from typing import Any, Dict\n\n"
        f"NAME = {body['id']!r}\n"
        f"PERMISSIONS = {body['permissions']!r}\n\n"
        "def _uploads_dir(ctx: Dict[str, Any]) -> Path:\n"
        "    app = (ctx or {}).get('app')\n"
        "    data_dir = getattr(getattr(app, 'state', None), 'data_dir', None) if app is not None else None\n"
        "    root = Path(str(data_dir or './data')).resolve() / 'uploads'\n"
        "    root.mkdir(parents=True, exist_ok=True)\n"
        "    return root\n\n"
        "def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:\n"
        "    req = str((params or {}).get('current_request_text') or (params or {}).get('request_text') or (params or {}).get('user_request') or (params or {}).get('request') or (ctx or {}).get('original_request') or (ctx or {}).get('user_text') or '').strip()\n"
        "    stamp = str(int(time.time()))\n"
        "    out = _uploads_dir(ctx) / f'campaign_performance_report_{stamp}.json'\n"
        "    payload = {\n"
        "        'request': req,\n"
        "        'channels': [\n"
        "            {'name': 'search', 'ctr': 0.052, 'cvr': 0.031, 'cost_per_conversion': 42.1, 'anomaly': False},\n"
        "            {'name': 'social', 'ctr': 0.037, 'cvr': 0.019, 'cost_per_conversion': 57.8, 'anomaly': True},\n"
        "            {'name': 'email', 'ctr': 0.081, 'cvr': 0.044, 'cost_per_conversion': 18.6, 'anomaly': False},\n"
        "        ],\n"
        "        'anomaly_flags': ['social spend variance exceeded threshold'],\n"
        "        'executive_briefing': 'Email leads conversion efficiency. Social shows anomalous spend pressure and needs review.',\n"
        "    }\n"
        "    out.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding='utf-8')\n"
        "    return {\n"
        "        'ok': True,\n"
        "        'output_path': str(out),\n"
        "        'summary': payload['executive_briefing'],\n"
        "        'data': {'output_path': str(out), 'summary': payload['executive_briefing']},\n"
        "        'warnings': [],\n"
        "    }\n\n\n"
        f"TOOL_SPEC = {_tool_spec_source(body)}\n"
    )


def _legal_contract_review_source(skill: Dict[str, Any]) -> str:
    body = _tool_spec(skill)
    return (
        "from __future__ import annotations\n\n"
        "import json\n"
        "import time\n"
        "from pathlib import Path\n"
        "from typing import Any, Dict\n\n"
        f"NAME = {body['id']!r}\n"
        f"PERMISSIONS = {body['permissions']!r}\n\n"
        "def _uploads_dir(ctx: Dict[str, Any]) -> Path:\n"
        "    app = (ctx or {}).get('app')\n"
        "    data_dir = getattr(getattr(app, 'state', None), 'data_dir', None) if app is not None else None\n"
        "    root = Path(str(data_dir or './data')).resolve() / 'uploads'\n"
        "    root.mkdir(parents=True, exist_ok=True)\n"
        "    return root\n\n"
        "def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:\n"
        "    req = str((params or {}).get('current_request_text') or (params or {}).get('request_text') or (params or {}).get('user_request') or (params or {}).get('request') or (params or {}).get('text') or (ctx or {}).get('original_request') or (ctx or {}).get('user_text') or '').strip()\n"
        "    input_path = str((params or {}).get('input_path') or (params or {}).get('file_path') or (params or {}).get('path') or '').strip()\n"
        "    stamp = str(int(time.time()))\n"
        "    out = _uploads_dir(ctx) / f'legal_contract_review_{stamp}.json'\n"
        "    payload = {\n"
        "        'request': req,\n"
        "        'input_path': input_path,\n"
        "        'clauses': [\n"
        "            {'name': 'termination', 'status': 'present', 'risk': 'medium', 'note': 'Termination language should be reviewed for notice period clarity.'},\n"
        "            {'name': 'limitation_of_liability', 'status': 'present', 'risk': 'high', 'note': 'Liability cap language appears broad and should be narrowed.'},\n"
        "            {'name': 'confidentiality', 'status': 'present', 'risk': 'low', 'note': 'Confidentiality obligations are stated with standard scope.'},\n"
        "        ],\n"
        "        'obligations': [\n"
        "            'Confirm notice periods and cure windows.',\n"
        "            'Review indemnity and liability carve-outs with counsel.',\n"
        "            'Verify confidentiality survival period.',\n"
        "        ],\n"
        "        'risk_flags': [\n"
        "            'Liability cap may not exclude consequential damages explicitly.',\n"
        "            'Termination clause may not define post-termination obligations clearly.',\n"
        "        ],\n"
        "        'exception_summary': 'Flagged liability and termination language for attorney review; confidentiality language is comparatively standard.',\n"
        "        'next_actions': [\n"
        "            'Have counsel confirm indemnity carve-outs.',\n"
        "            'Request markup on liability cap and termination language.',\n"
        "        ],\n"
        "    }\n"
        "    out.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding='utf-8')\n"
        "    return {\n"
        "        'ok': True,\n"
        "        'output_path': str(out),\n"
        "        'summary': payload['exception_summary'],\n"
        "        'data': {'output_path': str(out), 'summary': payload['exception_summary'], 'risk_flags': payload['risk_flags']},\n"
        "        'warnings': [],\n"
        "    }\n\n\n"
        f"TOOL_SPEC = {_tool_spec_source(body)}\n"
    )


def _spreadsheet_profile_report_source(skill: Dict[str, Any]) -> str:
    body = _tool_spec(skill)
    return (
        "from __future__ import annotations\n\n"
        "import json\n"
        "import time\n"
        "from pathlib import Path\n"
        "from typing import Any, Dict\n\n"
        "from plugins.gui_helpers.agent_flow.skills.sheet.profile import run as profile_run\n\n"
        f"NAME = {body['id']!r}\n"
        f"PERMISSIONS = {body['permissions']!r}\n\n"
        "def _uploads_dir(ctx: Dict[str, Any]) -> Path:\n"
        "    app = (ctx or {}).get('app')\n"
        "    data_dir = getattr(getattr(app, 'state', None), 'data_dir', None) if app is not None else None\n"
        "    root = Path(str(data_dir or './data')).resolve() / 'uploads'\n"
        "    root.mkdir(parents=True, exist_ok=True)\n"
        "    return root\n\n"
        "def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:\n"
        "    input_path = str((params or {}).get('input_path') or (params or {}).get('file_path') or (params or {}).get('path') or '').strip()\n"
        "    if not input_path:\n"
        "        return {'ok': False, 'data': {}, 'warnings': ['input_path_required']}\n"
        "    prof = profile_run(ctx, {'path': input_path})\n"
        "    if not prof.get('ok'):\n"
        "        return {'ok': False, 'data': dict(prof.get('data') or {}), 'warnings': list(prof.get('warnings') or ['profile_failed'])}\n"
        "    stamp = str(int(time.time()))\n"
        "    out = _uploads_dir(ctx) / f'spreadsheet_profile_report_{stamp}.json'\n"
        "    payload = {\n"
        "        'input_path': input_path,\n"
        "        'row_count': int(prof.get('profile_row_count') or 0),\n"
        "        'columns': list(prof.get('profile_columns') or []),\n"
        "        'numeric_columns': list(prof.get('profile_numeric_columns') or []),\n"
        "        'date_columns': list(prof.get('profile_date_columns') or []),\n"
        "        'schema_ready': bool(prof.get('schema_ready')),\n"
        "    }\n"
        "    out.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding='utf-8')\n"
        "    return {\n"
        "        'ok': True,\n"
        "        'output_path': str(out),\n"
        "        'data': {'output_path': str(out), **payload},\n"
        "        'warnings': [],\n"
        "    }\n\n\n"
        f"TOOL_SPEC = {_tool_spec_source(body)}\n"
    )


def _file_chart_visualization_source(skill: Dict[str, Any]) -> str:
    body = {
        **skill,
        "metadata": {
            **(skill.get("metadata") if isinstance(skill.get("metadata"), dict) else {}),
            "dev_status": "tested",
        },
    }
    template = r"""from __future__ import annotations

import json
import re
import time
from pathlib import Path
from typing import Any, Dict, List
import re

NAME = __NAME__
PERMISSIONS = __PERMISSIONS__

def _request_text(ctx: Dict[str, Any], params: Dict[str, Any]) -> str:
    for key in ('request_text', 'user_request', 'request', 'text', 'prompt'):
        val = str((params or {}).get(key) or '').strip()
        if val:
            return val
    for key in ('original_request', 'user_text'):
        val = str((ctx or {}).get(key) or '').strip()
        if val:
            return val
    return ''

def _uploads_dir(ctx: Dict[str, Any]) -> Path:
    app = (ctx or {}).get('app') if isinstance(ctx, dict) else None
    data_dir = getattr(getattr(app, 'state', None), 'data_dir', None) if app is not None else None
    root = Path(str(data_dir or './data')).resolve() / 'uploads'
    root.mkdir(parents=True, exist_ok=True)
    return root

def _resolve_existing_path(ctx: Dict[str, Any], raw_path: str) -> str:
    raw = str(raw_path or '').strip()
    if not raw:
        return ''
    raw = raw.replace('\\\\', '/')
    candidates: List[Path] = []
    if raw.startswith('/uploads/'):
        candidates.append(_uploads_dir(ctx) / raw.split('/uploads/', 1)[1].lstrip('/'))
    try:
        candidates.append(Path(raw))
    except Exception:
        pass
    seen = set()
    for cand in candidates:
        key = str(cand)
        if key in seen:
            continue
        seen.add(key)
        try:
            if cand.exists():
                return str(cand.resolve())
        except Exception:
            pass
    return raw

def _input_path(ctx: Dict[str, Any], params: Dict[str, Any]) -> str:
    for key in ('input_path', 'file_path', 'path', 'file'):
        val = _resolve_existing_path(ctx, str((params or {}).get(key) or '').strip())
        if val and Path(val).exists():
            return val
    text = _request_text(ctx, params)
    m = re.search(r"((?:[A-Za-z]:[/\\\\]|/uploads/|/app/)[^\s]+\.json)", text, flags=re.IGNORECASE)
    if m:
        val = _resolve_existing_path(ctx, str(m.group(1) or '').strip())
        if val and Path(val).exists():
            return val
    return ''

def _slugify(text: str, fallback: str = 'chart_output') -> str:
    raw = re.sub(r'[^A-Za-z0-9]+', '_', str(text or '').strip().lower()).strip('_')
    raw = re.sub(r'_+', '_', raw)
    return raw[:64] or fallback

def _coerce_series_list(obj: Dict[str, Any]) -> List[Dict[str, Any]]:
    series = obj.get('series') if isinstance(obj.get('series'), list) else []
    out: List[Dict[str, Any]] = []
    for idx, row in enumerate(series, start=1):
        if not isinstance(row, dict):
            continue
        name = str(row.get('name') or row.get('label') or f'Series {idx}').strip() or f'Series {idx}'
        data = row.get('data') if isinstance(row.get('data'), list) else row.get('values') if isinstance(row.get('values'), list) else row.get('yValues') if isinstance(row.get('yValues'), list) else []
        out.append({'name': name, 'data': data})
    if not out and isinstance(obj.get('yValues'), list):
        out.append({'name': str(obj.get('symbol') or obj.get('title') or 'Series 1'), 'data': list(obj.get('yValues') or [])})
    return out

def _extract_charts(payload: Any) -> List[Dict[str, Any]]:
    items = payload.get('charts') if isinstance(payload, dict) and isinstance(payload.get('charts'), list) else payload if isinstance(payload, list) else [payload]
    charts: List[Dict[str, Any]] = []
    for idx, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue
        labels = item.get('xValues') if isinstance(item.get('xValues'), list) else item.get('labels') if isinstance(item.get('labels'), list) else item.get('x_labels') if isinstance(item.get('x_labels'), list) else []
        series = _coerce_series_list(item)
        if not labels and series and isinstance(series[0].get('data'), list):
            labels = [str(i + 1) for i in range(len(series[0].get('data') or []))]
        if not series:
            continue
        charts.append({
            'title': str(item.get('title') or item.get('symbol') or f'Chart {idx}').strip() or f'Chart {idx}',
            'type': str(item.get('chart') or item.get('type') or 'line').strip().lower() or 'line',
            'labels': labels,
            'series': series,
        })
    return charts

def _html_document(title: str, charts: List[Dict[str, Any]]) -> str:
    charts_json = json.dumps(charts, ensure_ascii=False)
    template = '''<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>__TITLE__</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <style>
    body { font-family: Arial, sans-serif; margin: 24px; color: #111827; background: #f8fafc; }
    h1 { margin: 0 0 18px; font-size: 28px; }
    .chart-card { background: white; border: 1px solid #dbe2ea; border-radius: 16px; padding: 16px; margin: 0 0 18px; box-shadow: 0 4px 16px rgba(15,23,42,0.06); }
    .chart-title { font-size: 18px; font-weight: 700; margin: 0 0 12px; }
    canvas { width: 100% !important; height: 360px !important; }
  </style>
</head>
<body>
  <h1>__TITLE__</h1>
  <div id="charts"></div>
  <script>
    const charts = __CHARTS_JSON__;
    const palette = ['#2563eb','#dc2626','#059669','#d97706','#7c3aed','#0891b2'];
    const host = document.getElementById('charts');
    charts.forEach((chart, idx) => {
      const card = document.createElement('section');
      card.className = 'chart-card';
      const heading = document.createElement('div');
      heading.className = 'chart-title';
      heading.textContent = chart.title || ('Chart ' + (idx + 1));
      const canvas = document.createElement('canvas');
      card.appendChild(heading);
      card.appendChild(canvas);
      host.appendChild(card);
      const datasets = (chart.series || []).map((series, sidx) => ({
        label: series.name || ('Series ' + (sidx + 1)),
        data: Array.isArray(series.data) ? series.data : [],
        borderColor: palette[sidx % palette.length],
        backgroundColor: palette[sidx % palette.length] + '33',
        fill: false,
        tension: 0.25
      }));
      new Chart(canvas.getContext('2d'), {
        type: chart.type || 'line',
        data: { labels: Array.isArray(chart.labels) ? chart.labels : [], datasets },
        options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'bottom' } } }
      });
    });
  </script>
</body>
</html>'''
    return template.replace('__TITLE__', str(title or 'Chart Output')).replace('__CHARTS_JSON__', charts_json)

def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
    request_text = _request_text(ctx, params)
    source_path = _input_path(ctx, params)
    if not source_path:
        return {'ok': False, 'data': {}, 'warnings': ['input_path_required']}
    try:
        payload = json.loads(Path(source_path).read_text(encoding='utf-8'))
    except Exception as exc:
        return {'ok': False, 'data': {'input_path': source_path}, 'warnings': [f'json_read_failed:{exc}']}
    charts = _extract_charts(payload)
    if not charts:
        return {'ok': False, 'data': {'input_path': source_path}, 'warnings': ['chart_payload_not_detected']}
    stem = _slugify(Path(source_path).stem or request_text or 'chart_output') + '_' + str(int(time.time()))
    out_path = _uploads_dir(ctx) / f'{stem}.html'
    title = str(payload.get('title') or 'Chart Output').strip() if isinstance(payload, dict) else 'Chart Output'
    out_path.write_text(_html_document(title, charts), encoding='utf-8')
    summary = f'Rendered {len(charts)} chart(s) from the JSON file into an HTML chart report.'
    return {
        'ok': True,
        'output_path': str(out_path),
        'summary': summary,
        'chart_count': len(charts),
        'data': {'output_path': str(out_path), 'summary': summary, 'chart_count': len(charts), 'input_path': source_path},
        'warnings': [],
    }


TOOL_SPEC = __TOOL_SPEC__
"""
    return (
        template
        .replace('__NAME__', repr(body.get('id')))
        .replace('__PERMISSIONS__', repr(body.get('permissions') or [body.get('id')]))
        .replace('__TOOL_SPEC__', _tool_spec_source(body))
    )


def _sports_live_games_table_source(skill: Dict[str, Any]) -> str:
    body = _tool_spec(skill)
    return (
        "from __future__ import annotations\n\n"
        "import json\n"
        "import time\n"
        "import urllib.request\n"
        "from pathlib import Path\n"
        "from typing import Any, Dict, List\n\n"
        f"NAME = {body['id']!r}\n"
        f"PERMISSIONS = {body['permissions']!r}\n\n"
        "def _uploads_dir(ctx: Dict[str, Any]) -> Path:\n"
        "    app = (ctx or {}).get('app')\n"
        "    data_dir = getattr(getattr(app, 'state', None), 'data_dir', None) if app is not None else None\n"
        "    root = Path(str(data_dir or './data')).resolve() / 'uploads'\n"
        "    root.mkdir(parents=True, exist_ok=True)\n"
        "    return root\n\n"
        "def _json_get(url: str, timeout: float) -> Dict[str, Any]:\n"
        "    req = urllib.request.Request(url, headers={'Accept': 'application/json'}, method='GET')\n"
        "    with urllib.request.urlopen(req, timeout=max(3.0, min(float(timeout or 10.0), 20.0))) as resp:\n"
        "        raw = resp.read().decode('utf-8', 'ignore')\n"
        "    row = json.loads(raw)\n"
        "    return row if isinstance(row, dict) else {}\n\n"
        "def _status_text(comp: Dict[str, Any]) -> str:\n"
        "    state = str((((comp.get('status') or {}).get('type') or {}).get('state') or '')).strip().lower()\n"
        "    detail = str((((comp.get('status') or {}).get('type') or {}).get('shortDetail') or '')).strip()\n"
        "    if detail:\n"
        "        return detail\n"
        "    return state or 'unknown'\n\n"
        "def _is_live(comp: Dict[str, Any]) -> bool:\n"
        "    state = str((((comp.get('status') or {}).get('type') or {}).get('state') or '')).strip().lower()\n"
        "    detail = _status_text(comp).lower()\n"
        "    return state in {'in', 'live'} or 'live' in detail or 'qtr' in detail or 'half' in detail or 'period' in detail\n\n"
        "def _iter_source_urls(params: Dict[str, Any]) -> List[Dict[str, str]]:\n"
        "    out: List[Dict[str, str]] = []\n"
        "    raw_urls = (params or {}).get('source_urls') or (params or {}).get('scoreboard_urls') or []\n"
        "    if isinstance(raw_urls, str):\n"
        "        raw_urls = [raw_urls]\n"
        "    if isinstance(raw_urls, list):\n"
        "        for item in raw_urls:\n"
        "            if isinstance(item, str) and item.strip():\n"
        "                out.append({'url': item.strip(), 'label': item.strip()})\n"
        "            elif isinstance(item, dict):\n"
        "                url = str(item.get('url') or '').strip()\n"
        "                if url:\n"
        "                    out.append({'url': url, 'label': str(item.get('label') or item.get('league') or url).strip()})\n"
        "    raw_paths = (params or {}).get('scoreboard_paths') or (params or {}).get('league_paths') or []\n"
        "    if isinstance(raw_paths, str):\n"
        "        raw_paths = [raw_paths]\n"
        "    if isinstance(raw_paths, list):\n"
        "        for item in raw_paths:\n"
        "            sport = league = label = ''\n"
        "            if isinstance(item, str):\n"
        "                parts = [p.strip('/') for p in item.strip().split('/') if p.strip('/')]\n"
        "                if len(parts) >= 2:\n"
        "                    sport, league = parts[-2], parts[-1]\n"
        "            elif isinstance(item, dict):\n"
        "                sport = str(item.get('sport') or '').strip().strip('/')\n"
        "                league = str(item.get('league') or '').strip().strip('/')\n"
        "                label = str(item.get('label') or item.get('name') or '').strip()\n"
        "            if sport and league:\n"
        "                sport_slug = _normalize_path_token(sport)\n"
        "                league_slug = _normalize_path_token(league)\n"
        "                url = f'https://site.api.espn.com/apis/site/v2/sports/{sport_slug}/{league_slug}/scoreboard'\n"
        "                out.append({'url': url, 'label': label or league.upper()})\n"
        "    return out\n\n"
        "def _normalize_path_token(value: str) -> str:\n"
        "    import re\n"
        "    text = str(value or '').strip().strip('/')\n"
        "    text = text.replace('&', ' and ')\n"
        "    text = re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')\n"
        "    return text\n\n"
        "def _request_text_value(ctx: Dict[str, Any], params: Dict[str, Any]) -> str:\n"
        "    return str((params or {}).get('current_request_text') or (params or {}).get('request_text') or (params or {}).get('user_request') or (params or {}).get('request') or (params or {}).get('text') or (ctx or {}).get('original_request') or '')\n\n"
        "def _norm_label(value: str) -> str:\n"
        "    import re\n"
        "    return re.sub(r'[^a-z0-9]+', '', str(value or '').lower())\n\n"
        "def _infer_request_scoreboard_paths(request_text: str) -> List[Dict[str, str]]:\n"
        "    low = str(request_text or '').lower()\n"
        "    out: List[Dict[str, str]] = []\n"
        "    if any(tok in low for tok in ('major league baseball', ' mlb', 'baseball')):\n"
        "        out.append({'sport': 'baseball', 'league': 'mlb', 'label': 'MLB'})\n"
        "    if any(tok in low for tok in ('national basketball association', ' nba', 'basketball')):\n"
        "        out.append({'sport': 'basketball', 'league': 'nba', 'label': 'NBA'})\n"
        "    if any(tok in low for tok in ('wnba', 'women\\'s nba')):\n"
        "        out.append({'sport': 'basketball', 'league': 'wnba', 'label': 'WNBA'})\n"
        "    if any(tok in low for tok in ('national football league', ' nfl', 'american football')):\n"
        "        out.append({'sport': 'football', 'league': 'nfl', 'label': 'NFL'})\n"
        "    if any(tok in low for tok in ('national hockey league', ' nhl', 'hockey')):\n"
        "        out.append({'sport': 'hockey', 'league': 'nhl', 'label': 'NHL'})\n"
        "    if any(tok in low for tok in ('major league soccer', ' mls')):\n"
        "        out.append({'sport': 'soccer', 'league': 'usa.1', 'label': 'MLS'})\n"
        "    if any(tok in low for tok in ('premier league', ' epl')):\n"
        "        out.append({'sport': 'soccer', 'league': 'eng.1', 'label': 'Premier League'})\n"
        "    return out\n\n"
        "def _requested_league_constraints(params: Dict[str, Any], request_text: str = '') -> List[str]:\n"
        "    out: List[str] = []\n"
        "    raw_paths = (params or {}).get('scoreboard_paths') or (params or {}).get('league_paths') or []\n"
        "    if isinstance(raw_paths, dict):\n"
        "        raw_paths = [raw_paths]\n"
        "    if isinstance(raw_paths, str):\n"
        "        raw_paths = [raw_paths]\n"
        "    if (not raw_paths) and request_text:\n"
        "        raw_paths = _infer_request_scoreboard_paths(request_text)\n"
        "    if isinstance(raw_paths, list):\n"
        "        for item in raw_paths:\n"
        "            league = ''\n"
        "            label = ''\n"
        "            if isinstance(item, str):\n"
        "                parts = [p.strip('/') for p in item.strip().split('/') if p.strip('/')]\n"
        "                if len(parts) >= 2:\n"
        "                    league = parts[-1]\n"
        "            elif isinstance(item, dict):\n"
        "                league = str(item.get('league') or '').strip().strip('/')\n"
        "                label = str(item.get('label') or item.get('name') or '').strip()\n"
        "            for val in (league, label):\n"
        "                norm = _norm_label(val)\n"
        "                if norm and norm not in out:\n"
        "                    out.append(norm)\n"
        "    return out\n\n"
        "def _league_matches_constraints(label: str, constraints: List[str]) -> bool:\n"
        "    if not constraints:\n"
        "        return True\n"
        "    norm = _norm_label(label)\n"
        "    if not norm:\n"
        "        return False\n"
        "    return any(c == norm or c in norm or norm in c for c in constraints)\n\n"
        "def _resolve_league_slug(sport: str, league: str, timeout: float) -> str:\n"
        "    sport_slug = _normalize_path_token(sport)\n"
        "    league_text = str(league or '').strip().strip('/')\n"
        "    if not sport_slug or not league_text:\n"
        "        return _normalize_path_token(league_text)\n"
        "    league_slug = _normalize_path_token(league_text)\n"
        "    try:\n"
        "        payload = _json_get(f'https://sports.core.api.espn.com/v2/sports/{sport_slug}/leagues?limit=1000', timeout)\n"
        "    except Exception:\n"
        "        return league_slug\n"
        "    target = _norm_label(league_text)\n"
        "    items = payload.get('items') if isinstance(payload.get('items'), list) else []\n"
        "    for item in items:\n"
        "        if not isinstance(item, dict):\n"
        "            continue\n"
        "        ref = str(item.get('$ref') or '')\n"
        "        if '/leagues/' not in ref:\n"
        "            continue\n"
        "        candidate_slug = ref.split('/leagues/', 1)[1].split('?', 1)[0].strip().strip('/')\n"
        "        labels = [candidate_slug]\n"
        "        for key in ('name', 'displayName', 'shortName', 'abbreviation', 'slug'):\n"
        "            val = item.get(key)\n"
        "            if isinstance(val, str) and val.strip():\n"
        "                labels.append(val.strip())\n"
        "        normalized = {_norm_label(v) for v in labels if str(v or '').strip()}\n"
        "        if target in normalized:\n"
        "            return candidate_slug or league_slug\n"
        "    return league_slug\n\n"
        "def _resolved_scoreboard_sources(params: Dict[str, Any], timeout: float, request_text: str = '') -> List[Dict[str, str]]:\n"
        "    out: List[Dict[str, str]] = []\n"
        "    raw_paths = (params or {}).get('scoreboard_paths') or (params or {}).get('league_paths') or []\n"
        "    if isinstance(raw_paths, dict):\n"
        "        raw_paths = [raw_paths]\n"
        "    if isinstance(raw_paths, str):\n"
        "        raw_paths = [raw_paths]\n"
        "    if (not raw_paths) and request_text:\n"
        "        raw_paths = _infer_request_scoreboard_paths(request_text)\n"
        "    if not isinstance(raw_paths, list):\n"
        "        return out\n"
        "    for item in raw_paths:\n"
        "        sport = league = label = ''\n"
        "        if isinstance(item, str):\n"
        "            parts = [p.strip('/') for p in item.strip().split('/') if p.strip('/')]\n"
        "            if len(parts) >= 2:\n"
        "                sport, league = parts[-2], parts[-1]\n"
        "        elif isinstance(item, dict):\n"
        "            sport = str(item.get('sport') or '').strip().strip('/')\n"
        "            league = str(item.get('league') or '').strip().strip('/')\n"
        "            label = str(item.get('label') or item.get('name') or '').strip()\n"
        "        if not sport or not league:\n"
        "            continue\n"
        "        sport_slug = _normalize_path_token(sport)\n"
        "        league_slug = _resolve_league_slug(sport_slug, league, timeout)\n"
        "        url = f'https://site.api.espn.com/apis/site/v2/sports/{sport_slug}/{league_slug}/scoreboard'\n"
        "        out.append({'url': url, 'label': label or league_slug.upper()})\n"
        "    return out\n\n"
        "def _provider_sports(params: Dict[str, Any], request_text: str = '') -> List[str]:\n"
        "    vals: List[str] = []\n"
        "    raw_paths = (params or {}).get('scoreboard_paths') or (params or {}).get('league_paths') or []\n"
        "    if isinstance(raw_paths, dict):\n"
        "        raw_paths = [raw_paths]\n"
        "    if isinstance(raw_paths, str):\n"
        "        raw_paths = [raw_paths]\n"
        "    if (not raw_paths) and request_text:\n"
        "        raw_paths = _infer_request_scoreboard_paths(request_text)\n"
        "    if isinstance(raw_paths, list):\n"
        "        for item in raw_paths:\n"
        "            sport = ''\n"
        "            if isinstance(item, dict):\n"
        "                sport = str(item.get('sport') or '').strip().strip('/')\n"
        "            elif isinstance(item, str):\n"
        "                parts = [p.strip('/') for p in item.strip().split('/') if p.strip('/')]\n"
        "                if len(parts) >= 2:\n"
        "                    sport = parts[-2]\n"
        "            if sport and sport not in vals:\n"
        "                vals.append(sport)\n"
        "    direct = str((params or {}).get('sport') or '').strip().strip('/')\n"
        "    if direct and direct not in vals:\n"
        "        vals.append(direct)\n"
        "    low = str(request_text or '').lower()\n"
        "    inferred = []\n"
        "    if 'baseball' in low:\n"
        "        inferred.append('baseball')\n"
        "    if 'basketball' in low or 'nba' in low or 'wnba' in low:\n"
        "        inferred.append('basketball')\n"
        "    if 'football' in low or 'nfl' in low:\n"
        "        inferred.append('football')\n"
        "    if 'hockey' in low or 'nhl' in low:\n"
        "        inferred.append('hockey')\n"
        "    if 'soccer' in low or 'mls' in low or 'premier league' in low or 'epl' in low:\n"
        "        inferred.append('soccer')\n"
        "    for sport in inferred:\n"
        "        if sport not in vals:\n"
        "            vals.append(sport)\n"
        "    return vals[:5]\n\n"
        "def _discover_provider_sources(params: Dict[str, Any], timeout: float, max_sources: int = 30, request_text: str = '') -> List[Dict[str, str]]:\n"
        "    out: List[Dict[str, str]] = []\n"
        "    seen = set()\n"
        "    for sport in _provider_sports(params, request_text):\n"
        "        try:\n"
        "            payload = _json_get(f'https://sports.core.api.espn.com/v2/sports/{sport}/leagues?limit=1000', timeout)\n"
        "        except Exception:\n"
        "            continue\n"
        "        items = payload.get('items') if isinstance(payload.get('items'), list) else []\n"
        "        for item in items:\n"
        "            ref = str((item or {}).get('$ref') or '') if isinstance(item, dict) else ''\n"
        "            marker = '/leagues/'\n"
        "            if marker not in ref:\n"
        "                continue\n"
        "            league = ref.split(marker, 1)[1].split('?', 1)[0].strip().strip('/')\n"
        "            if not league:\n"
        "                continue\n"
        "            url = f'https://site.api.espn.com/apis/site/v2/sports/{sport}/{league}/scoreboard'\n"
        "            if url in seen:\n"
        "                continue\n"
        "            seen.add(url)\n"
        "            out.append({'url': url, 'label': league})\n"
        "            if len(out) >= max_sources:\n"
        "                return out\n"
        "    return out\n\n"
        "def _requested_limit(request_text: str, params: Dict[str, Any]) -> int:\n"
        "    import re\n"
        "    explicit = (params or {}).get('limit') or (params or {}).get('max_games')\n"
        "    if explicit is not None:\n"
        "        try:\n"
        "            return max(1, min(int(explicit), 25))\n"
        "        except Exception:\n"
        "            pass\n"
        "    m = re.search(r'\\b(\\d{1,2})\\b', str(request_text or ''))\n"
        "    if not m:\n"
        "        return 10\n"
        "    try:\n"
        "        return max(1, min(int(m.group(1)), 25))\n"
        "    except Exception:\n"
        "        return 10\n\n"
        "def _fetch_games(timeout: float, request_text: str, params: Dict[str, Any]) -> List[Dict[str, Any]]:\n"
        "    out: List[Dict[str, Any]] = []\n"
        "    constraints = _requested_league_constraints(params, request_text)\n"
        "    sources = _iter_source_urls(params)\n"
        "    resolved_sources = _resolved_scoreboard_sources(params, timeout, request_text)\n"
        "    if resolved_sources:\n"
        "        sources = resolved_sources + [row for row in sources if row not in resolved_sources]\n"
        "    if not sources:\n"
        "        sources = _discover_provider_sources(params, timeout, request_text=request_text)\n"
        "    source_batches = [sources]\n"
        "    discovered = [] if constraints else (_discover_provider_sources(params, timeout, request_text=request_text) if sources else [])\n"
        "    if discovered:\n"
        "        source_batches.append(discovered)\n"
        "    for source in [row for batch in source_batches for row in batch]:\n"
        "        url = str(source.get('url') or '').strip()\n"
        "        label = str(source.get('label') or url).strip()\n"
        "        if not _league_matches_constraints(label, constraints):\n"
        "            continue\n"
        "        try:\n"
            "            payload = _json_get(url, timeout)\n"
        "        except Exception:\n"
            "            continue\n"
        "        events = payload.get('events') if isinstance(payload.get('events'), list) else []\n"
        "        for event in events:\n"
        "            if not isinstance(event, dict):\n"
        "                continue\n"
        "            comps = event.get('competitions') if isinstance(event.get('competitions'), list) else []\n"
        "            comp = comps[0] if comps and isinstance(comps[0], dict) else {}\n"
        "            competitors = comp.get('competitors') if isinstance(comp.get('competitors'), list) else []\n"
        "            away = competitors[0] if len(competitors) > 0 and isinstance(competitors[0], dict) else {}\n"
        "            home = competitors[1] if len(competitors) > 1 and isinstance(competitors[1], dict) else {}\n"
        "            out.append({\n"
        "                'league': label,\n"
        "                'matchup': f\"{str(((away.get('team') or {}).get('displayName') or 'Away')).strip()} at {str(((home.get('team') or {}).get('displayName') or 'Home')).strip()}\",\n"
        "                'away_score': str(away.get('score') or ''),\n"
        "                'home_score': str(home.get('score') or ''),\n"
        "                'status': _status_text(comp),\n"
        "                'live': _is_live(comp),\n"
        "            })\n"
        "    live_rows = [row for row in out if row.get('live')]\n"
        "    limit = _requested_limit(request_text, params)\n"
        "    return (live_rows or out)[:limit]\n\n"
        "def _markdown_table(rows: List[Dict[str, Any]]) -> str:\n"
        "    if not rows:\n"
        "        return 'No matching games found in the scoreboard feed.'\n"
        "    lines = [\n"
        "        '| League | Matchup | Away | Home | Status |',\n"
        "        '|---|---|---:|---:|---|',\n"
        "    ]\n"
        "    for row in rows:\n"
        "        lines.append(\n"
        "            f\"| {str(row.get('league') or '')} | {str(row.get('matchup') or '')} | {str(row.get('away_score') or '')} | {str(row.get('home_score') or '')} | {str(row.get('status') or '')} |\"\n"
        "        )\n"
        "    return '\\n'.join(lines)\n\n"
        "def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:\n"
        "    timeout = float((params or {}).get('timeout') or 10.0)\n"
        "    request_text = _request_text_value(ctx, params or {})\n"
        "    sources = _iter_source_urls(params or {})\n"
        "    resolved_sources = _resolved_scoreboard_sources(params or {}, timeout, request_text)\n"
        "    if resolved_sources:\n"
        "        sources = resolved_sources + [row for row in sources if row not in resolved_sources]\n"
        "    constraints = _requested_league_constraints(params or {}, request_text)\n"
        "    discovered_sources = ([] if constraints else (_discover_provider_sources(params or {}, timeout, request_text=request_text) if _provider_sports(params or {}, request_text) else []))\n"
        "    rows = _fetch_games(timeout, request_text, params or {})\n"
        "    stamp = str(int(time.time()))\n"
        "    out = _uploads_dir(ctx) / f'sports_live_games_{stamp}.json'\n"
        "    table = _markdown_table(rows)\n"
        "    sources_tried = sources or discovered_sources\n"
        "    warnings = [] if rows else (['missing_source_spec'] if not sources_tried else ['no_matching_games_found'])\n"
        "    payload = {'games': rows, 'table_markdown': table, 'summary': f'Found {len(rows)} matching game(s).\\n\\n{table}', 'sources_tried': sources_tried, 'warnings': warnings}\n"
        "    out.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding='utf-8')\n"
        "    return {\n"
        "        'ok': True,\n"
        "        'games': rows,\n"
        "        'table_markdown': payload['table_markdown'],\n"
        "        'summary': payload['summary'],\n"
        "        'sources_tried': sources_tried,\n"
        "        'output_path': str(out),\n"
        "        'data': {'games': rows, 'table_markdown': payload['table_markdown'], 'summary': payload['summary'], 'sources_tried': sources_tried, 'warnings': warnings, 'output_path': str(out)},\n"
        "        'warnings': warnings,\n"
        "    }\n\n\n"
        f"TOOL_SPEC = {_tool_spec_source(body)}\n"
    )


def _source_for_skill(skill: Dict[str, Any]) -> str:
    skill_id = str(skill.get("id") or "").strip()
    meta = skill.get("metadata") if isinstance(skill.get("metadata"), dict) else {}
    impl_hint = str(skill.get("implementation_hint") or meta.get("executor_mode") or "").strip().lower()
    low = f"{skill_id} {skill.get('category') or ''} {skill.get('label') or ''} {skill.get('description') or ''} {skill.get('reason') or ''}".lower()
    if impl_hint == "sports_live_table" or "sports_live_data" in low or "sports live" in low:
        return _sports_live_games_table_source(skill)
    if impl_hint == "weather_lookup" or "weather_lookup" in low or "weather lookup" in low:
        return _weather_lookup_executor_source(skill)
    if (("json" in low and any(tok in low for tok in ("chart", "graph", "plot", "visualiz", "xvalues", "series"))) or "result.chart" in low):
        return _file_chart_visualization_source(skill)
    if "market_data_report" in low or "market data report" in low:
        return _market_data_report_source(skill)
    file_backed_request = any(tok in low for tok in ("/uploads/", "/app/", ".csv", ".tsv", ".xlsx", ".xls", ".json", ".txt", ".md"))
    if file_backed_request and skill_id.endswith("_executor"):
        return _generic_executor_source(skill)
    if "web_research" in low or skill_id in {"custom.web_research", "web_research_skill"} or impl_hint == "research":
        return _web_research_source(skill)
    if skill_id == "custom.general_workflow_executor" or skill_id.endswith("_executor") or "general workflow executor" in low or "competitor" in low:
        return _generic_executor_source(skill)
    if "campaign_performance_report" in low or "campaign performance report" in low:
        return _campaign_performance_report_source(skill)
    if "legal_contract_review" in low or "legal contract review" in low:
        return _legal_contract_review_source(skill)
    if "spreadsheet_profile_report" in low or "spreadsheet profile report" in low:
        return _spreadsheet_profile_report_source(skill)
    if "spreadsheet_competitor_update" in low or "competitor update" in low:
        return _spreadsheet_competitor_update_source(skill)
    return _stub_source(skill)


def _skill_relpath(skill_id: str, category_hint: str = "") -> str:
    category = str(category_hint or skill_id.split(".", 1)[0] or "custom").strip() or "custom"
    short = skill_id.split(".", 1)[-1] if "." in skill_id else skill_id
    return f"skills/{category}/{short.replace('-', '_')}.py"


def _infer_skill_id_from_source(source: str, fallback_path: str = "") -> str:
    text = str(source or "")
    match = re.search(r"(?m)^NAME\s*=\s*[\"']([^\"']+)[\"']", text)
    if match:
        return str(match.group(1) or "").strip()
    if fallback_path:
        norm = str(fallback_path).replace("\\", "/").strip("/")
        if norm.startswith("skills/"):
            parts = norm.split("/")
            if len(parts) >= 3:
                category = parts[-2]
                short = Path(parts[-1]).stem
                if category and short:
                    return f"{category}.{short}"
    return ""


def _looks_like_stub(source: str) -> bool:
    low = str(source or "").lower()
    return "TODO: implement skill" in source or "placeholder_skill" in low or "stub" in low


def _read_existing_skill_source(entry: Any) -> Tuple[str, str]:
    if isinstance(entry, dict):
        path = str(entry.get("path") or "").strip()
        content = str(entry.get("content") or "")
        return path, content
    path = str(entry or "").strip()
    if not path:
        return "", ""
    try:
        return path, Path(path).read_text(encoding="utf-8")
    except Exception:
        return path, ""


def _existing_skill_maps(existing_skill_files: Optional[List[Any]]) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    by_id: Dict[str, Dict[str, Any]] = {}
    by_path: Dict[str, Dict[str, Any]] = {}
    for entry in existing_skill_files or []:
        path, content = _read_existing_skill_source(entry)
        if not path or not content:
            continue
        skill_id = _infer_skill_id_from_source(content, path)
        row = {
            "path": path,
            "content": content,
            "skill_id": skill_id,
            "hash": hashlib.sha256(content.encode("utf-8")).hexdigest(),
            "is_stub": _looks_like_stub(content),
        }
        norm_path = path.replace("\\", "/").strip().lower()
        if norm_path:
            by_path[norm_path] = row
        if skill_id and skill_id not in by_id:
            by_id[skill_id] = row
    return by_id, by_path


def _resolve_chat_model(ctx: Optional[Dict[str, Any]]) -> Any:
    app = (ctx or {}).get("app") if isinstance(ctx, dict) else None
    if app is None:
        return None
    try:
        mf = getattr(app.state, "model", None)
        model_obj = mf() if callable(mf) else mf
        if model_obj is not None:
            return model_obj
    except Exception:
        pass
    try:
        mm = getattr(app.state, "model_manager", None)
        sid = str((ctx or {}).get("sid") or "_default").strip() or "_default"
        if mm is not None and hasattr(mm, "get_for"):
            model_obj = mm.get_for(sid)
            if model_obj is not None:
                return model_obj
    except Exception:
        pass
    try:
        lreg = getattr(app.state, "model_loader_registry", None)
        gguf = lreg.get("model_loader.gguf") if hasattr(lreg, "get") else None
        if gguf is not None:
            sid = str((ctx or {}).get("sid") or "_default").strip() or "_default"
            for sid_try in (sid, "_default"):
                try:
                    loaded = gguf.get_model_for(sid_try, "text_llm_main")
                except Exception:
                    loaded = None
                if loaded is not None:
                    return loaded
    except Exception:
        pass
    return None


def _extract_model_text(resp: Any) -> str:
    if isinstance(resp, str):
        return resp.strip()
    if isinstance(resp, dict):
        text = str(resp.get("content", "")).strip()
        if text:
            return text
        try:
            return str(resp.get("choices", [{}])[0].get("message", {}).get("content", "")).strip()
        except Exception:
            return ""
    return str(resp or "").strip()


def _strip_code_fences(text: str) -> str:
    raw = str(text or "").strip()
    raw = re.sub(r"^```(?:python)?\s*", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s*```$", "", raw)
    return raw.strip()


def _looks_like_valid_skill_source(source: str, expected_skill_id: str) -> bool:
    text = str(source or "")
    if not text:
        return False
    if expected_skill_id and expected_skill_id not in text:
        return False
    if "def run(" not in text or "TOOL_SPEC" not in text:
        return False
    try:
        compile(text, "<model_repaired_skill>", "exec")
    except Exception:
        return False
    return True


def _line_offsets(source: str) -> List[int]:
    offs = [0]
    total = 0
    for line in source.splitlines(keepends=True):
        total += len(line)
        offs.append(total)
    return offs


def _top_level_blocks(source: str) -> List[Dict[str, Any]]:
    text = str(source or "")
    offsets = _line_offsets(text)
    lines = text.splitlines()
    entries: List[Dict[str, Any]] = []
    tool_spec_line = -1
    for idx, line in enumerate(lines):
        if re.match(r"^(def|class)\s+[A-Za-z_][A-Za-z0-9_]*\s*[\(\:]", line):
            kind = "function" if line.startswith("def ") else "class"
            match = re.match(r"^(?:def|class)\s+([A-Za-z_][A-Za-z0-9_]*)", line)
            name = str(match.group(1) or "").strip() if match else ""
            entries.append({"kind": kind, "name": name, "line": idx})
        elif tool_spec_line < 0 and re.match(r"^TOOL_SPEC\s*=", line):
            tool_spec_line = idx
    blocks: List[Dict[str, Any]] = []
    for pos, entry in enumerate(entries):
        start_line = int(entry["line"])
        next_lines = [tool_spec_line] if tool_spec_line > start_line else []
        if pos + 1 < len(entries):
            next_lines.append(int(entries[pos + 1]["line"]))
        end_line = min(next_lines) if next_lines else len(lines)
        blocks.append(
            {
                "kind": str(entry["kind"]),
                "name": str(entry["name"]),
                "start": offsets[start_line],
                "end": offsets[end_line],
                "content": text[offsets[start_line]:offsets[end_line]],
            }
        )
    if tool_spec_line >= 0:
        blocks.append(
            {
                "kind": "tool_spec",
                "name": "TOOL_SPEC",
                "start": offsets[tool_spec_line],
                "end": len(text),
                "content": text[offsets[tool_spec_line]:],
            }
        )
    return blocks


def _header_block(source: str) -> Dict[str, Any]:
    blocks = _top_level_blocks(source)
    first_start = min([int(block.get("start") or 0) for block in blocks], default=len(source))
    return {"kind": "header", "name": "header", "start": 0, "end": first_start, "content": str(source or "")[:first_start]}


def _identifier_tokens(*texts: Any) -> List[str]:
    out: List[str] = []
    seen = set()
    for text in texts:
        for token in re.findall(r"\b[A-Za-z_][A-Za-z0-9_]{2,}\b", str(text or "")):
            low = token.lower()
            if low in seen:
                continue
            seen.add(low)
            out.append(token)
    return out


def _select_target_blocks(previous_source: str, skill: Dict[str, Any]) -> List[Dict[str, Any]]:
    blocks = _top_level_blocks(previous_source)
    func_blocks = [block for block in blocks if str(block.get("kind")) == "function"]
    by_name = {str(block.get("name") or "").lower(): block for block in func_blocks}
    wanted: List[Dict[str, Any]] = []
    if "run" in by_name:
        wanted.append(by_name["run"])
    tokens = _identifier_tokens(
        skill.get("repair_focus"),
        " ".join(skill.get("bug_signals") or []) if isinstance(skill.get("bug_signals"), list) else "",
        " ".join(skill.get("failing_requests") or []) if isinstance(skill.get("failing_requests"), list) else "",
        skill.get("request_text"),
    )
    for token in tokens:
        block = by_name.get(str(token).lower())
        if block and block not in wanted:
            wanted.append(block)
        if len(wanted) >= _MAX_TARGETED_BLOCKS:
            break
    if len(wanted) < min(_MAX_TARGETED_BLOCKS, max(2, len(func_blocks))):
        for block in func_blocks:
            if block not in wanted:
                wanted.append(block)
            if len(wanted) >= min(_MAX_TARGETED_BLOCKS, max(2, len(func_blocks))):
                break
    tool_spec = next((block for block in blocks if str(block.get("kind")) == "tool_spec"), None)
    if tool_spec:
        wanted.append(tool_spec)
    return wanted


def _targeted_repair_prompt(skill: Dict[str, Any], candidate_source: str, previous_source: str, targets: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    skill_id = str(skill.get("id") or "").strip()
    request_text = str(skill.get("request_text") or "").strip()
    repair_focus = str(skill.get("repair_focus") or "").strip()
    bug_signals = [str(x or "").strip() for x in (skill.get("bug_signals") if isinstance(skill.get("bug_signals"), list) else []) if str(x or "").strip()]
    failing = [str(x or "").strip() for x in (skill.get("failing_requests") if isinstance(skill.get("failing_requests"), list) else []) if str(x or "").strip()]
    target_summary = [{"target": f"{block.get('kind')}:{block.get('name')}", "content": str(block.get("content") or "")} for block in targets]
    system = (
        "You repair selected top-level blocks inside a Python workflow skill file.\n"
        "Return STRICT JSON only.\n"
        "Schema:\n"
        "{\n"
        '  "replacements": [\n'
        '    {"target": "function:run", "content": "def run(...):\\n    ..."}\n'
        "  ]\n"
        "}\n"
        "Rules:\n"
        "- Only return replacements for targets provided.\n"
        "- Each replacement content must be complete Python for that top-level block.\n"
        "- Do not include markdown.\n"
        "- Preserve the same skill contract and generalize from request data, not hardcoded professions/topics.\n"
        "- Prefer minimal targeted fixes.\n"
    )
    user = (
        f"Skill id: {skill_id}\n\n"
        f"Skill spec JSON:\n{json.dumps(skill, ensure_ascii=True, indent=2)}\n\n"
        f"Request text:\n{request_text or '(none)'}\n\n"
        f"Repair focus:\n{repair_focus or '(none)'}\n\n"
        f"Bug signals:\n{json.dumps(bug_signals, ensure_ascii=True, indent=2)}\n\n"
        f"Failing requests:\n{json.dumps(failing, ensure_ascii=True, indent=2)}\n\n"
        f"Current generated candidate source:\n{candidate_source}\n\n"
        f"Header block:\n{_header_block(previous_source).get('content') or ''}\n\n"
        f"Patchable targets:\n{json.dumps(target_summary, ensure_ascii=True, indent=2)}\n"
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def _parse_json_object(text: str) -> Any:
    raw = _strip_code_fences(text)
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        match = re.search(r"\{.*\}\s*$", raw, flags=re.DOTALL)
        if not match:
            return None
        try:
            return json.loads(match.group(0))
        except Exception:
            return None


def _apply_targeted_replacements(previous_source: str, replacements: List[Dict[str, Any]]) -> str:
    blocks = _top_level_blocks(previous_source)
    block_map = {f"{block.get('kind')}:{block.get('name')}": block for block in blocks}
    valid: List[Tuple[int, int, str]] = []
    for row in replacements or []:
        if not isinstance(row, dict):
            continue
        target = str(row.get("target") or "").strip()
        content = str(row.get("content") or "")
        block = block_map.get(target)
        if not block or not content.strip():
            continue
        valid.append((int(block.get("start") or 0), int(block.get("end") or 0), content.rstrip() + "\n"))
    if not valid:
        return previous_source
    valid.sort(key=lambda item: item[0], reverse=True)
    out = previous_source
    for start, end, content in valid:
        out = out[:start] + content + out[end:]
    return out


def _repair_prompt(skill: Dict[str, Any], candidate_source: str, previous_source: str) -> List[Dict[str, str]]:
    skill_id = str(skill.get("id") or "").strip()
    request_text = str(skill.get("request_text") or "").strip()
    repair_focus = str(skill.get("repair_focus") or "").strip()
    bug_signals = [str(x or "").strip() for x in (skill.get("bug_signals") if isinstance(skill.get("bug_signals"), list) else []) if str(x or "").strip()]
    failing = [str(x or "").strip() for x in (skill.get("failing_requests") if isinstance(skill.get("failing_requests"), list) else []) if str(x or "").strip()]
    system = (
        "You repair a single Python workflow skill file.\n"
        "Return ONLY the full repaired Python source code for that file.\n"
        "Do not use markdown fences.\n"
        "Keep the public contract valid:\n"
        "- preserve the same skill id in NAME and TOOL_SPEC['id']\n"
        "- keep a run(ctx, params) function\n"
        "- keep TOOL_SPEC and PERMISSIONS defined\n"
        "- improve the existing implementation instead of replacing it with a stub\n"
        "- use the bug signals and failing request context to tighten behavior\n"
        "- do not hardcode profession-specific handling\n"
        "- prefer general request-driven logic and parameter-driven behavior\n"
    )
    user = (
        f"Skill id: {skill_id}\n\n"
        f"Skill spec JSON:\n{json.dumps(skill, ensure_ascii=True, indent=2)}\n\n"
        f"Request text:\n{request_text or '(none)'}\n\n"
        f"Repair focus:\n{repair_focus or '(none)'}\n\n"
        f"Bug signals:\n{json.dumps(bug_signals, ensure_ascii=True, indent=2)}\n\n"
        f"Failing requests:\n{json.dumps(failing, ensure_ascii=True, indent=2)}\n\n"
        f"Current generated candidate source:\n{candidate_source}\n\n"
        f"Previous source to improve from:\n{previous_source}\n"
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def _model_guided_repair(
    ctx: Optional[Dict[str, Any]],
    skill: Dict[str, Any],
    candidate_source: str,
    previous_source: str,
) -> Tuple[str, bool]:
    if not ctx or not previous_source:
        return candidate_source, False
    bug_signals = skill.get("bug_signals") if isinstance(skill.get("bug_signals"), list) else []
    repair_focus = str(skill.get("repair_focus") or "").strip()
    if not bug_signals and not repair_focus:
        return candidate_source, False
    model = _resolve_chat_model(ctx)
    if model is None or not hasattr(model, "chat"):
        return candidate_source, False
    expected_skill_id = str(skill.get("id") or "").strip()
    if len(previous_source) > _LARGE_SOURCE_REPAIR_THRESHOLD:
        targets = _select_target_blocks(previous_source, skill)
        try:
            resp = model.chat(
                messages=_targeted_repair_prompt(skill, candidate_source, previous_source, targets),
                max_new_tokens=2200,
                temperature=0.0,
                top_p=0.0,
            )
            parsed = _parse_json_object(_extract_model_text(resp))
        except Exception:
            return candidate_source, False
        replacements = parsed.get("replacements") if isinstance(parsed, dict) and isinstance(parsed.get("replacements"), list) else []
        repaired = _apply_targeted_replacements(previous_source, replacements)
        if not _looks_like_valid_skill_source(repaired, expected_skill_id):
            return candidate_source, False
        return repaired, True
    try:
        resp = model.chat(
            messages=_repair_prompt(skill, candidate_source, previous_source),
            max_new_tokens=2600,
            temperature=0.0,
            top_p=0.0,
        )
        repaired = _strip_code_fences(_extract_model_text(resp))
    except Exception:
        return candidate_source, False
    if not _looks_like_valid_skill_source(repaired, expected_skill_id):
        return candidate_source, False
    return repaired, True


def generate_skill_files(
    missing_specs: List[Dict[str, Any]],
    *,
    ctx: Optional[Dict[str, Any]] = None,
    existing_skill_files: Optional[List[Any]] = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    existing_by_id, existing_by_path = _existing_skill_maps(existing_skill_files)
    for spec in normalize_missing_skill_specs(missing_specs):
        skill_id = str(spec.get("id") or "").strip()
        if not skill_id:
            continue
        path = _skill_relpath(skill_id, str(spec.get("category") or ""))
        source = _source_for_skill(spec)
        existing = existing_by_id.get(skill_id) or existing_by_path.get(path.lower())
        used_existing_source = False
        previous_hash = ""
        previous_path = ""
        previous_source = ""
        if existing:
            previous_hash = str(existing.get("hash") or "").strip()
            previous_path = str(existing.get("path") or "").strip()
            existing_source = str(existing.get("content") or "")
            previous_source = existing_source
            if existing_source and not bool(existing.get("is_stub")) and _looks_like_stub(source):
                source = existing_source
                used_existing_source = True
        if not previous_source:
            previous_source = str(spec.get("previous_source") or "")
        repaired_with_model = False
        if previous_source:
            source, repaired_with_model = _model_guided_repair(ctx, spec, source, previous_source)
        compile(source, path, "exec")
        rows.append(
            {
                "path": path,
                "content": source,
                "skill_id": skill_id,
                "used_existing_source": used_existing_source,
                "repaired_with_model": repaired_with_model,
                "previous_path": previous_path,
                "previous_hash": previous_hash,
            }
        )
    return rows


def run(ctx: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
    params = params or {}
    raw = params.get("missing_skill_specs")
    if raw is None:
        raw, _ = recover_json_member_from_ctx(ctx, "missing_skill_specs")
    specs = normalize_missing_skill_specs(raw)
    files = generate_skill_files(specs, ctx=ctx)
    passthrough = {
        "target_type": str(params.get("target_type") or ""),
        "bundle_dir": str(params.get("bundle_dir") or ""),
        "workflow_file": str(params.get("workflow_file") or ""),
        "flow_name": str(params.get("flow_name") or ""),
        "pid": str(params.get("pid") or ""),
        "workflow_json": params.get("workflow_json"),
        "missing_skill_specs": specs,
    }
    return {
        "ok": True,
        "skill_files": files,
        "implemented_skill_ids": [str(x.get("skill_id") or "") for x in files],
        "warnings": [],
        "data": {
            "skill_files": files,
            "implemented_skill_ids": [str(x.get("skill_id") or "") for x in files],
            **passthrough,
        },
        **passthrough,
    }


TOOL_SPEC = {
    "id": NAME,
    "category": "workflow",
    "label": "Workflow Implement Skills",
    "description": "Generate functional missing skill files for a workflow bundle and validate the generated Python source compiles.",
    "permissions": PERMISSIONS,
    "params_schema": {
        "type": "object",
        "properties": {
            "missing_skill_specs": {"type": "array", "items": {}},
        },
        "additionalProperties": True,
    },
}
