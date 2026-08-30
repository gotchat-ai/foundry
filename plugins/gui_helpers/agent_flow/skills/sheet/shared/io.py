import csv
import json
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional, Sequence

from shared.utils import as_path, ensure_parent, ext_of, infer_delimiter, read_text_preview, unique_headers, row_to_dict, dict_to_row


def _resolve_input_path(file: str) -> Path:
    """
    Resolve spreadsheet input paths with lightweight repo-aware fallback.
    Supports bare filenames staged under common runtime folders such as
    data/agent_workflow/repo.
    """
    path = as_path(file)
    if path.exists():
        return path
    raw = str(file or "").strip()
    if not raw:
        return path
    raw_path = Path(raw)
    name = raw_path.name
    cwd = Path(os.getcwd()).resolve()
    if raw.startswith('/uploads/'):
        upload_name = raw.split('/uploads/', 1)[-1].lstrip('/')
        upload_roots = [
            cwd / 'uploads',
            cwd / 'llmloader2' / 'uploads',
            cwd / 'data' / 'uploads',
            cwd / 'llmloader2' / 'data' / 'uploads',
        ]
        for parent in list(cwd.parents)[:3]:
            upload_roots.extend([
                parent / 'uploads',
                parent / 'llmloader2' / 'uploads',
                parent / 'data' / 'uploads',
                parent / 'llmloader2' / 'data' / 'uploads',
            ])
        seen_uploads = set()
        for root0 in upload_roots:
            key = str(root0)
            if key in seen_uploads:
                continue
            seen_uploads.add(key)
            cand = (root0 / upload_name).resolve()
            if cand.exists():
                return cand
    roots = [
        cwd,
        cwd / "llmloader2",
        cwd / "data",
        cwd / "llmloader2" / "data",
        cwd / "data" / "uploads",
        cwd / "llmloader2" / "data" / "uploads",
        cwd / "data" / "agent_workflow" / "repo",
        cwd / "llmloader2" / "data" / "agent_workflow" / "repo",
        cwd / "generated",
        cwd / "llmloader2" / "generated",
    ]
    # Also probe a couple of parent roots in case service CWD is nested.
    for p in list(cwd.parents)[:3]:
        roots.extend(
            [
                p,
                p / "llmloader2",
                p / "data",
                p / "llmloader2" / "data",
                p / "data" / "uploads",
                p / "llmloader2" / "data" / "uploads",
                p / "data" / "agent_workflow" / "repo",
                p / "llmloader2" / "data" / "agent_workflow" / "repo",
                p / "generated",
                p / "llmloader2" / "generated",
            ]
        )
    # De-duplicate while preserving order.
    seen_roots = set()
    uniq_roots = []
    for r in roots:
        rs = str(r)
        if rs in seen_roots:
            continue
        seen_roots.add(rs)
        uniq_roots.append(r)
    roots = uniq_roots

    # 1) Try exact raw path under each candidate root.
    for root in roots:
        cand = (root / raw).resolve()
        if cand.exists():
            return cand
    # 2) If caller supplied a nested/guessed path, retry basename directly in roots.
    if str(raw_path).replace("\\", "/").find("/") >= 0 and name:
        for root in roots:
            cand = (root / name).resolve()
            if cand.exists():
                return cand
    # 2b) If the model omitted the spreadsheet extension, resolve common spreadsheet suffixes.
    if name and not Path(name).suffix:
        exts = (".xlsx", ".xlsm", ".xls", ".csv", ".tsv", ".ods")
        for root in roots:
            for ext in exts:
                cand = (root / f"{raw}{ext}").resolve()
                if cand.exists():
                    return cand
                cand = (root / f"{name}{ext}").resolve()
                if cand.exists():
                    return cand
        for root in roots:
            if not root.exists():
                continue
            for ext in exts:
                try:
                    hits = list(root.rglob(f"{name}{ext}"))
                except Exception:
                    hits = []
                for hit in hits:
                    try:
                        if hit.is_file():
                            return hit.resolve()
                    except Exception:
                        continue
    # 3) Recursive basename search fallback.
    for root in roots:
        if not root.exists():
            continue
        try:
            hits = list(root.rglob(name))
        except Exception:
            hits = []
        for hit in hits:
            try:
                if hit.is_file():
                    return hit.resolve()
            except Exception:
                continue
    return path


def _looks_like_number(text: str) -> bool:
    s = str(text or "").strip()
    if not s:
        return False
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s))


def _header_score(row: Sequence[object]) -> int:
    vals = [str(v).strip() for v in list(row or [])]
    non_empty = [v for v in vals if v]
    if not non_empty:
        return -10
    alpha_like = sum(1 for v in non_empty if any(ch.isalpha() for ch in v))
    numeric_like = sum(1 for v in non_empty if _looks_like_number(v))
    unique_nonempty = len(set(non_empty))
    return (alpha_like * 3) + unique_nonempty - (numeric_like * 2)


def _pick_header_index(rows: Sequence[Sequence[object]]) -> int:
    best_idx, best_score = 0, -10**9
    for i, row in enumerate(rows):
        score = _header_score(row)
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _normalize_header_name(name: str) -> str:
    return "".join(ch for ch in str(name or "").lower() if ch.isalnum())


def _likely_key_columns(headers: Sequence[str]) -> list[str]:
    if not headers:
        return []
    wanted = {
        "orderno",
        "orderid",
        "id",
        "customer",
        "customername",
        "customerid",
        "name",
    }
    out = []
    for h in headers:
        nh = _normalize_header_name(h)
        if nh in wanted or any(w in nh for w in ("order", "customer")):
            out.append(str(h))
    return out


def _row_is_data_record(headers: Sequence[str], row: Sequence[object]) -> bool:
    vals = ["" if v is None else str(v).strip() for v in list(row or [])]
    non_empty = sum(1 for v in vals if v)
    if non_empty == 0:
        return False
    rec = row_to_dict(headers, row)
    key_cols = _likely_key_columns(headers)
    if key_cols:
        key_vals = [str(rec.get(k) or "").strip() for k in key_cols]
        if not any(key_vals):
            return False
    # Ignore obvious metadata/tax rows that appear above/below tabular data.
    joined = " ".join(v.lower() for v in vals if v)
    if joined in {"tax", "total", "subtotal"}:
        return False
    return True

def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except Exception as e:
        raise RuntimeError("openpyxl is required for xlsx/xlsm support. Install with: pip install openpyxl") from e

def _iter_xlsx(path: Path, sheet: Optional[str] = None, data_only: bool = True):
    try:
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        probe = []
        try:
            for _ in range(25):
                probe.append(next(rows))
        except StopIteration:
            pass
        if not probe:
            wb.close()
            return
        header_i = _pick_header_index(probe)
        headers = unique_headers(probe[header_i])
        row_no = header_i + 2
        for row in probe[header_i + 1:]:
            if _row_is_data_record(headers, row):
                yield headers, list(row), row_no
            row_no += 1
        for row in rows:
            if _row_is_data_record(headers, row):
                yield headers, list(row), row_no
            row_no += 1
        wb.close()
        return
    except Exception:
        # Fallback parser for .xlsx when openpyxl is unavailable.
        pass

    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"

    def _col_idx(cell_ref: str) -> int:
        letters = "".join(ch for ch in str(cell_ref or "") if ch.isalpha())
        n = 0
        for ch in letters:
            n = n * 26 + (ord(ch.upper()) - 64)
        return max(0, n - 1)

    with zipfile.ZipFile(path, "r") as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(f"{ns}si"):
                shared.append("".join((t.text or "") for t in si.findall(".//" + ns + "t")))

        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = wb.find(f"{ns}sheets")
        if sheets is None or not list(sheets):
            return

        sheet_name_map: dict[str, str] = {}
        for sh in sheets:
            sh_name = sh.attrib.get("name") or ""
            rel_id = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id") or ""
            if rel_id:
                sheet_name_map[sh_name] = rel_id
        chosen_rel_id = None
        if sheet and sheet in sheet_name_map:
            chosen_rel_id = sheet_name_map[sheet]
        else:
            first = list(sheets)[0]
            chosen_rel_id = first.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if not chosen_rel_id:
            return

        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        target = None
        for rel in rels:
            if rel.attrib.get("Id") == chosen_rel_id:
                target = rel.attrib.get("Target")
                break
        if not target:
            return
        sheet_path = "xl/" + str(target).lstrip("/")
        sh = ET.fromstring(zf.read(sheet_path))
        data = sh.find(f"{ns}sheetData")
        if data is None:
            return

        parsed_rows: list[list[object]] = []
        for row in data.findall(f"{ns}row"):
            vals: dict[int, object] = {}
            for c in row.findall(f"{ns}c"):
                idx = _col_idx(c.attrib.get("r", "A1"))
                ctype = c.attrib.get("t")
                v_el = c.find(f"{ns}v")
                v: object = ""
                if ctype == "s" and v_el is not None and v_el.text is not None:
                    try:
                        si = int(v_el.text)
                        v = shared[si] if 0 <= si < len(shared) else ""
                    except Exception:
                        v = ""
                elif ctype == "inlineStr":
                    is_el = c.find(f"{ns}is")
                    if is_el is not None:
                        v = "".join((t.text or "") for t in is_el.findall(".//" + ns + "t"))
                elif v_el is not None and v_el.text is not None:
                    v = v_el.text
                vals[idx] = v
            if vals:
                max_i = max(vals.keys())
                arr = [""] * (max_i + 1)
                for k, v in vals.items():
                    arr[k] = v
                parsed_rows.append(arr)

        if not parsed_rows:
            return
        probe = parsed_rows[:25]
        header_i = _pick_header_index(probe)
        headers = unique_headers(parsed_rows[header_i])
        for i, row in enumerate(parsed_rows[header_i + 1 :], start=header_i + 2):
            if _row_is_data_record(headers, row):
                yield headers, list(row), i

def _iter_csv(path: Path, delimiter: Optional[str] = None):
    sample = read_text_preview(path)
    delim = delimiter or infer_delimiter(path, sample)
    with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as f:
        reader = csv.reader(f, delimiter=delim)
        probe = []
        try:
            for _ in range(25):
                probe.append(next(reader))
        except StopIteration:
            pass
        if not probe:
            return
        header_i = _pick_header_index(probe)
        headers = unique_headers(probe[header_i])
        row_no = header_i + 2
        for row in probe[header_i + 1:]:
            yield headers, row, row_no
            row_no += 1
        for row in reader:
            yield headers, row, row_no
            row_no += 1

def _iter_ods(path: Path, sheet: Optional[str] = None):
    try:
        from odf.opendocument import load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
    except Exception as e:
        raise RuntimeError("ODS support requires odfpy. Install with: pip install odfpy") from e
    def cell_text(cell):
        vals = [str(p) for p in cell.getElementsByType(P)]
        text = "".join(vals)
        return text if text != "" else cell.getAttribute("value") or cell.getAttribute("datevalue") or ""
    doc = load(str(path))
    tables = doc.spreadsheet.getElementsByType(Table)
    table = None
    if sheet:
        for t in tables:
            if t.getAttribute("name") == sheet:
                table = t
                break
    table = table or (tables[0] if tables else None)
    if table is None:
        return
    raw_rows = []
    for tr in table.getElementsByType(TableRow):
        vals = []
        for cell in tr.getElementsByType(TableCell):
            repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
            vals.extend([cell_text(cell)] * repeat)
        raw_rows.append(vals)
    if not raw_rows:
        return
    header_i = _pick_header_index(raw_rows[:25])
    headers = unique_headers(raw_rows[header_i])
    for idx, row in enumerate(raw_rows[header_i + 1:], start=header_i + 2):
        yield headers, row, idx

def _iter_xls_optional(path: Path, sheet: Optional[str] = None):
    try:
        import pandas as pd
    except Exception as e:
        raise RuntimeError("Legacy .xls support requires optional dependencies: pip install pandas xlrd") from e
    df = pd.read_excel(path, sheet_name=sheet or 0, dtype=object)
    headers = unique_headers(list(df.columns))
    for pos, row in enumerate(df.itertuples(index=False, name=None), start=2):
        yield headers, list(row), pos

def iter_records(file: str, sheet: Optional[str] = None, offset: int = 0, limit: Optional[int] = None, columns: Optional[Sequence[str]] = None, delimiter: Optional[str] = None):
    path = _resolve_input_path(file)
    ext = ext_of(path)
    if ext in {".xlsx", ".xlsm"}:
        gen = _iter_xlsx(path, sheet=sheet)
    elif ext in {".csv", ".tsv"}:
        gen = _iter_csv(path, delimiter=delimiter)
    elif ext == ".ods":
        gen = _iter_ods(path, sheet=sheet)
    elif ext == ".xls":
        gen = _iter_xls_optional(path, sheet=sheet)
    else:
        raise ValueError(f"Unsupported spreadsheet extension for read: {ext}")
    selected = set(columns or [])
    yielded = 0
    skipped = 0
    for headers, row, idx in gen:
        if skipped < offset:
            skipped += 1
            continue
        record = row_to_dict(headers, row)
        if selected:
            record = {c: record.get(c) for c in columns}
        yield record
        yielded += 1
        if limit is not None and yielded >= limit:
            break

def workbook_metadata(file: str):
    path = _resolve_input_path(file)
    ext = ext_of(path)
    if ext in {".xlsx", ".xlsm"}:
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = [{"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column} for ws in wb.worksheets]
        wb.close()
        return {"file": str(path), "format": ext, "sheets": sheets}
    rows = 0
    headers = None
    for rec in iter_records(str(path), limit=None):
        if headers is None:
            headers = list(rec.keys())
        rows += 1
    return {"file": str(path), "format": ext, "sheets": [{"name": path.stem, "row_count": rows, "columns": headers or []}]}

def write_records(records, output: str, sheet_name: str = "Sheet1", columns: Optional[Sequence[str]] = None, format: Optional[str] = None):
    out = as_path(output)
    ensure_parent(out)
    ext = (format.lower() if format else ext_of(out)).lower()
    if not ext.startswith("."):
        ext = "." + ext
    if columns:
        headers = list(columns)
    else:
        headers, seen = [], set()
        for rec in records:
            for k in rec.keys():
                if k not in seen:
                    headers.append(k)
                    seen.add(k)
    if ext in {".xlsx", ".xlsm"}:
        openpyxl = _require_openpyxl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (sheet_name or "Sheet1")[:31]
        ws.append(headers)
        for rec in records:
            ws.append(dict_to_row(headers, rec))
        try:
            from openpyxl.styles import Font, PatternFill
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            for col_cells in ws.columns:
                letter = col_cells[0].column_letter
                max_len = 10
                for cell in list(col_cells)[:200]:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(val) + 2, 42))
                ws.column_dimensions[letter].width = max_len
        except Exception:
            pass
        wb.save(out)
        return {"ok": True, "output": str(out), "format": ext, "rows": len(records), "columns": headers}
    if ext in {".csv", ".tsv"}:
        delimiter = "\t" if ext == ".tsv" else ","
        with out.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter, extrasaction="ignore")
            writer.writeheader()
            for rec in records:
                writer.writerow({h: rec.get(h) for h in headers})
        return {"ok": True, "output": str(out), "format": ext, "rows": len(records), "columns": headers}
    if ext == ".json":
        with out.open("w", encoding="utf-8") as f:
            json.dump(list(records), f, indent=2, ensure_ascii=False, default=str)
        return {"ok": True, "output": str(out), "format": ext, "rows": len(records), "columns": headers}
    raise ValueError(f"Unsupported spreadsheet extension for write: {ext}")
